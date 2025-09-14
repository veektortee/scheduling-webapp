#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scheduler Testcase Builder GUI
- Merged format: { constants, run, calendar, shifts, providers }
- Tabs:
    1) Run (output dir, run controls, live log)
    2) Calendar (weekend days, month generator, calendar day list)
    3) Shifts (per-day shift table, add/update/delete; add across ALL days)
    4) Providers (edit provider, OFF dates (fixed/prefer), ON dates (fixed/prefer) + shift-type picker, prefs summary)
    5) Config (solver constants / weights / objective editors)
- Fixed: naming collision between Calendar and Providers day listboxes.
  * Calendar uses self.lst_days_cal
  * Providers now uses two panes: self.lst_days_off and self.lst_days_on
- Clear prefs on selected days works and updates summary/log.
- Color-coded OFF/ON day lists with legends.
"""
from __future__ import annotations

import json
import os
import sys
import threading
import subprocess
import queue
import random
from pathlib import Path
from datetime import date, timedelta, datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

import argparse, json, os, re, sys, subprocess, traceback
import datetime as dt
from collections import defaultdict
from typing import Dict, Any, List

import logging
from logging import Logger

from ortools.sat.python import cp_model
from openpyxl import Workbook
CHOSPITAL = ""
SCALE = 1
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
diagnose_schedule_api.py  — colorized, expanded diagnostics (multi-sheet aware)

Programmatic usage (no CLI):
  from diagnose_schedule_api import run
  run(case_path, schedule_path, no_color=False, preview=8)

Behavior is identical to the CLI version you shared:
- If the schedule is an Excel workbook (.xlsx/.xlsm), ALL suitable sheets
  (those having ShiftID and Provider/Assignee columns) are diagnosed.
  One report file is written per sheet next to the schedule file:
     <schedule_stem>__<SheetName>.diagnose.txt
- CSV/JSON behave as before (single report). For single-item inputs, prints to stdout.

Deps:
  - colorama (optional for terminal colors)
  - openpyxl (only if reading .xlsx/.xlsm)
"""

import json, csv, sys, collections, io
from pathlib import Path
from datetime import date, datetime

# -------------- Colors --------------
_USE_COLOR = True
try:
    from colorama import init as _colorama_init, Fore, Style
    _colorama_init()
except Exception:
    class _Dummy:
        def __getattr__(self, k): return ""
    Fore = Style = _Dummy()

def _c_ok(s):   return (Fore.GREEN + s + Style.RESET_ALL) if _USE_COLOR else s
def _c_fail(s): return (Fore.RED + s + Style.RESET_ALL)   if _USE_COLOR else s
def _c_warn(s): return (Fore.YELLOW + s + Style.RESET_ALL)if _USE_COLOR else s
def _c_head(s): return (Fore.CYAN + s + Style.RESET_ALL)  if _USE_COLOR else s
def _c_dim(s):  return (Style.DIM + s + Style.RESET_ALL)  if _USE_COLOR else s

# -------------- Robust readers --------------

_PREFERRED_ENCODINGS = ("utf-8", "utf-8-sig", "cp1252", "latin-1")

def _read_text_best_effort(path: str):
    for enc in _PREFERRED_ENCODINGS:
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                return f.read(), enc, False
        except UnicodeDecodeError:
            continue
    data = Path(path).read_bytes()
    txt = data.decode("latin-1", errors="replace")
    return txt, "latin-1-replace", True

def _read_json_best_effort(path: str):
    last_err = None
    for enc in _PREFERRED_ENCODINGS:
        try:
            with open(path, "r", encoding=enc) as f:
                return json.load(f), enc
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except json.JSONDecodeError:
            raise
    raw = Path(path).read_bytes().decode("latin-1", errors="replace")
    return json.loads(raw), "latin-1-replace"

# -------------- Case loader --------------

def load_case(path):
    case, enc = _read_json_best_effort(path)
    cal = case.get("calendar", {}) or {}
    days = list(cal.get("days", []))
    weekend_days = cal.get("weekend_days", ["Saturday","Sunday"])
    shifts = case.get("shifts", []) or []
    providers = case.get("providers", []) or []
    for p in providers:
        p.setdefault("name", p.get("id", ""))
        p.setdefault("type", "MD")
        p.setdefault("forbidden_days_hard", [])
        p.setdefault("forbidden_days_soft", [])
        p.setdefault("preferred_days_hard", {})
        p.setdefault("preferred_days_soft", {})
        lim = p.setdefault("limits", {})
        lim.setdefault("min_total", 0)
        lim.setdefault("max_total", None)
        lim.setdefault("type_ranges", {})
        if not isinstance(p.get("max_consecutive_days", None), int):
            p["max_consecutive_days"] = None
        if isinstance(p["preferred_days_hard"], list):
            p["preferred_days_hard"] = {}
        if isinstance(p["preferred_days_soft"], list):
            p["preferred_days_soft"] = {}
    return {
        "calendar_days": days,
        "weekend_days": weekend_days,
        "shifts": shifts,
        "providers": providers,
        "case_encoding": enc
    }

# -------------- Schedule loaders --------------

def infer_schedule_from_json(obj):
    if isinstance(obj, dict) and "assignments" in obj and isinstance(obj["assignments"], list):
        out = {}
        for row in obj["assignments"]:
            sid = row.get("shift_id") or row.get("id") or row.get("shift")
            prov = row.get("provider") or row.get("provider_name") or row.get("name")
            if sid and prov:
                out.setdefault(str(sid).strip(), []).append(str(prov).strip())
        return out
    if isinstance(obj, dict):
        ok = True
        for k in obj.keys():
            if not isinstance(k, str):
                ok = False; break
        if ok:
            out = {}
            for k, v in obj.items():
                if isinstance(v, list):
                    out[str(k).strip()] = [str(x).strip() for x in v if str(x).strip()]
                elif isinstance(v, str):
                    out[str(k).strip()] = [v.strip()]
            if out:
                return out
    return None

def _find_candidate_columns(header_lower):
    def find_col(names):
        for i, h in enumerate(header_lower):
            if h in names:
                return i
        return None
    c_shift = find_col({"shiftid", "shift id", "id", "shift"})
    c_prov  = find_col({"assignee", "provider", "provider_name", "name"})
    return c_shift, c_prov

def _load_schedules_from_xlsx(path: str):
    """Return list of (sheet_name, schedule_map). One entry per suitable sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to read .xlsx/.xlsm files. Install via: pip install openpyxl")

    wb = load_workbook(path, data_only=True, read_only=True)
    results = []

    for ws in wb.worksheets:
        # sniff header
        header = None
        for row in ws.iter_rows(max_row=10, values_only=True):
            if not row: continue
            nonempty = [c for c in row if c not in (None, "")]
            if len(nonempty) < 2: continue
            header = [str(c).strip() if c is not None else "" for c in row]
            break
        if not header:
            continue
        lower = [h.lower() for h in header]
        c_shift, c_prov = _find_candidate_columns(lower)
        if c_shift is None or c_prov is None:
            continue

        # read rows
        first = True
        sheet_map = {}
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row:
                continue
            sid = row[c_shift] if c_shift < len(row) else None
            prov = row[c_prov] if c_prov < len(row) else None
            if sid is None or str(sid).strip() == "":
                continue
            if prov is None or str(prov).strip() == "" or str(prov).strip().upper() == "UNFILLED":
                continue
            sid = str(sid).strip()
            prov = str(prov).strip()
            sheet_map.setdefault(sid, []).append(prov)

        if sheet_map:
            results.append((ws.title, sheet_map))

    if not results:
        sheet_names = ", ".join(ws.title for ws in wb.worksheets)
        raise ValueError(
            f"No suitable worksheets in {path}. Need headers: ShiftID & Provider/Assignee. "
            f"Sheets present: [{sheet_names}]"
        )
    return results

def _load_schedule_from_csv(path: str):
    text, enc, had_rep = _read_text_best_effort(str(path))
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        class _D(csv.excel):
            delimiter = ','
        dialect = _D()
    try:
        has_header = csv.Sniffer().has_header(sample)
    except Exception:
        has_header = True
    fobj = io.StringIO(text)
    reader = csv.DictReader(fobj, dialect=dialect) if has_header else csv.DictReader(fobj, fieldnames=None, dialect=dialect)
    col_shift = col_provider = None
    if reader.fieldnames:
        lower = [c.lower() for c in reader.fieldnames]
        c_shift, c_prov = _find_candidate_columns(lower)
        if c_shift is not None: col_shift = reader.fieldnames[c_shift]
        if c_prov  is not None: col_provider = reader.fieldnames[c_prov]
    if not col_shift or not col_provider:
        raise ValueError(f"CSV must include columns for shift and provider; got fields: {reader.fieldnames} (encoding={enc})")
    out = {}
    for row in reader:
        sid = (row.get(col_shift) or "").strip()
        prov = (row.get(col_provider) or "").strip()
        if sid and prov and prov.upper() != "UNFILLED":
            out.setdefault(sid, []).append(prov)
    return out, enc

def load_schedules(path_str: str):
    """
    Returns (items, source_descr)
    items is a list of (label, schedule_map).
      - For xlsx/xlsm: one per suitable sheet.
      - For json/csv: single item [(stem, schedule_map)].
    """
    p = Path(path_str)
    ext = p.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        sheets = _load_schedules_from_xlsx(str(p))
        return sheets, "xlsx"
    if ext == ".json":
        obj, enc = _read_json_best_effort(str(p))
        m = infer_schedule_from_json(obj)
        if m is None:
            raise ValueError("Unrecognized JSON schedule format. Expected {'assignments': [...]} or {shift_id: provider}.")
        return [(p.stem, m)], f"json({enc})"
    # CSV (or tsv-like)
    m, enc = _load_schedule_from_csv(str(p))
    return [(p.stem, m)], f"csv({enc})"

# -------------- Helpers --------------

def iso_weekday_name(dstr):
    y,m,dd = map(int, dstr.split("-"))
    return date(y,m,dd).strftime("%A")

def _to_date(dstr: str) -> date:
    y,m,dd = map(int, dstr.split("-"))
    return date(y,m,dd)

def _cluster_sizes(worked_days):
    if not worked_days:
        return []
    dd = sorted(_to_date(d) for d in set(worked_days))
    sizes = []; run = 1
    for i in range(1, len(dd)):
        if (dd[i] - dd[i-1]).days == 1:
            run += 1
        else:
            sizes.append(run); run = 1
    sizes.append(run)
    return sizes

# -------------- Diagnosis --------------

def diagnose(case, schedule_map, stream=sys.stdout, preview_limit=8, banner=None):
    days = case["calendar_days"]
    shifts = case["shifts"]
    providers = case["providers"]
    weekend_names = set(case["weekend_days"] or ["Saturday","Sunday"])

    if banner:
        print(_c_head(banner), file=stream)

    # Indexes
    shifts_by_id = {sh["id"]: sh for sh in shifts if "id" in sh}
    providers_by_name = {p["name"]: p for p in providers if "name" in p}
    day_to_shifts = collections.defaultdict(list)
    for s in shifts:
        day_to_shifts[s["date"]].append(s["id"])

    shift_date = {sid: sh["date"] for sid, sh in shifts_by_id.items()}
    # FIXED: correct dict comprehension (remove bad/duplicate line)
    shift_type = {sid: sh.get("type","") for sid, sh in shifts_by_id.items()}
    shift_allowed_types = {sid: set(sh.get("allowed_provider_types", ["MD"])) for sid, sh in shifts_by_id.items()}

    prov_day_to_shifts = collections.defaultdict(lambda: collections.defaultdict(list))
    for sid, provs in schedule_map.items():
        if sid not in shifts_by_id:
            continue
        d = shift_date[sid]
        for prov in provs:
            prov_day_to_shifts[prov][d].append(sid)

    checks = []
    def add_check(name, ok, details=""):
        checks.append((name, bool(ok), details))

    # 1) All shifts filled exactly once
    unfilled, overfilled = [], []
    for sid in shifts_by_id:
        n = len(schedule_map.get(sid, []))
        if n == 0:
            unfilled.append(sid)
        elif n > 1:
            overfilled.append((sid, n))
    add_check("All shifts filled exactly once", len(unfilled)==0 and len(overfilled)==0,
              (f"Unfilled:{len(unfilled)}; " if unfilled else "") + (f"Overfilled:{len(overfilled)}" if overfilled else ""))

    # 2) Unknown shift IDs
    unknown_shift_ids = [sid for sid in schedule_map if sid not in shifts_by_id]
    add_check("No unknown shift IDs in schedule", len(unknown_shift_ids)==0, f"Unknown IDs: {len(unknown_shift_ids)}")

    # 3) Provider exists
    unknown_providers = [prov for prov in prov_day_to_shifts.keys() if prov not in providers_by_name]
    add_check("All providers exist", len(unknown_providers)==0,
              f"Unknown providers: {', '.join(unknown_providers[:5])}{'...' if len(unknown_providers)>5 else ''}")

    # 4) Provider type allowed by shift
    bad_allowed = []
    for prov, by_day in prov_day_to_shifts.items():
        p = providers_by_name.get(prov)
        if not p: continue
        ptype = p.get("type","MD")
        for d, sids in by_day.items():
            for sid in sids:
                allowed = shift_allowed_types.get(sid) or set()
                if allowed and ptype not in allowed:
                    bad_allowed.append((prov, sid, ptype, sorted(allowed)))
    add_check("Provider type allowed for each assigned shift", len(bad_allowed)==0, f"Violations: {len(bad_allowed)}")

    # 5) Forbidden (hard-off) days
    bad_forbidden = []
    for prov, by_day in prov_day_to_shifts.items():
        p = providers_by_name.get(prov)
        if not p: continue
        forb = set(p.get("forbidden_days_hard", []))
        for d in by_day:
            if d in forb:
                bad_forbidden.append((prov, d, by_day[d]))
    add_check("Providers NOT scheduled on forbidden (hard-off) days", len(bad_forbidden)==0, f"Violations: {len(bad_forbidden)}")

    # 6) At most one shift per provider per day
