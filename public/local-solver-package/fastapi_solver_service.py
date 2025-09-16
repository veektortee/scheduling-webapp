#!/usr/bin/env python3
"""
FastAPI-based Solver Service for Medical Staff Scheduling

This high-performance service runs locally and provides WebSocket + REST API
for the Vercel-hosted web application to trigger optimization solver.

Features:
- FastAPI for maximum performance
- WebSocket for real-time progress updates
- Integrated OR-Tools solver logic from testcase_gui.py
- Background task processing for long-running optimizations
- Comprehensive error handling and logging

Usage:
    pip install fastapi uvicorn websockets python-multipart
    python fastapi_solver_service.py
    
The service will run on http://localhost:8000
"""

import json
import logging
import os
import sys
import asyncio
import uuid
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Any, List, Optional
import traceback
from concurrent.futures import ThreadPoolExecutor
import threading
import calendar as pycalendar
import shutil
from ortools.sat.python import cp_model
import tempfile

try:
    from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, BackgroundTasks
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.responses import JSONResponse, FileResponse
    from pydantic import BaseModel
    import uvicorn
    import testcase_gui as original_solver 
except ImportError:
    print("Please install required packages:")
    print("pip install fastapi uvicorn websockets python-multipart")
    sys.exit(1)

# Import the solver logic from your existing code
try:
    import testcase_gui as original_solver
    HAVE_ORIGINAL_SOLVER = True
    print("✅ Successfully imported testcase_gui.py as the solver engine.")
except ImportError:
    HAVE_ORIGINAL_SOLVER = False
    print("❌ WARNING: testcase_gui.py not found. Local solver will not work.")


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(name)s | %(message)s'
)
logger = logging.getLogger("scheduler-fastapi")

app = FastAPI(
    title="Medical Staff Scheduling Solver API",
    description="High-performance optimization service for medical staff scheduling",
    version="2.0.0"
)

# CORS setup for Vercel integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify your Vercel domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic models for API
class SchedulingCase(BaseModel):
    constants: Dict[str, Any]
    calendar: Dict[str, Any]
    shifts: List[Dict[str, Any]]
    providers: List[Dict[str, Any]]
    run: Optional[Dict[str, Any]] = None

class SolverStatus(BaseModel):
    status: str
    message: str
    run_id: Optional[str] = None
    progress: Optional[float] = None
    results: Optional[Dict[str, Any]] = None

# Global state management
active_runs: Dict[str, Dict[str, Any]] = {}
websocket_connections: Dict[str, WebSocket] = {}
thread_pool = ThreadPoolExecutor(max_workers=4)

class AdvancedSchedulingSolver:
    def __init__(self):
        self.output_dir = Path("solver_output")
        self.output_dir.mkdir(exist_ok=True)
        
    async def solve_async(self, case_data: Dict[str, Any], run_id: str) -> Dict[str, Any]:
        """
        Asynchronous wrapper for the solver that integrates your OR-Tools logic
        """
        loop = asyncio.get_event_loop()
        
        # Run the CPU-intensive solver in a thread pool
        result = await loop.run_in_executor(
            thread_pool, 
            self._solve_with_ortools, 
            case_data, 
            run_id
        )
        
        return result
    
    def _solve_with_ortools(self, case_data: Dict[str, Any], run_id: str) -> Dict[str, Any]:
        """
        Integrated OR-Tools solver logic adapted from your testcase_gui.py
        """
        try:
            # Prefer a deterministic output folder when caller provided run.out
            requested_out = None
            try:
                requested_out = case_data.get('run', {}).get('out')
            except Exception:
                requested_out = None

            if isinstance(requested_out, str) and requested_out.startswith('Result_'):
                # If the client supplies an explicit Result_N folder (e.g. Result_14)
                # prefer that deterministic name so uploads and downstream tooling
                # can rely on a stable path. We sanitize with os.path.basename to
                # avoid absolute paths or path traversal ("../"). If no valid
                # Result_* is provided we fall back to a run-id based folder.
                out_name = os.path.basename(requested_out)
                run_output_dir = self.output_dir / out_name
            else:
                run_output_dir = self.output_dir / run_id

            run_output_dir.mkdir(parents=True, exist_ok=True)
            
            # Save input case
            case_file = run_output_dir / "input_case.json"
            with open(case_file, 'w') as f:
                json.dump(case_data, f, indent=2)
            
            # Update progress
            self._update_progress(run_id, 10, "Initializing solver...")
            
            # Extract case components
            constants = case_data.get('constants', {})
            calendar_data = case_data.get('calendar', {})
            # Sanitize calendar days to avoid invalid dates arriving at the solver
            try:
                calendar_data = self._sanitize_calendar(calendar_data)
                # propagate back to case_data so saved input_case.json matches what solver sees
                case_data['calendar'] = calendar_data
            except Exception as e:
                logger.warning(f"Calendar sanitization failed: {e}")
            shifts = case_data.get('shifts', [])
            providers = case_data.get('providers', [])
            run_config = case_data.get('run', {})

            # Ensure that shift dates are present in the calendar. Some payloads
            # contain shifts referencing dates not included in calendar.days
            # (causes KeyError in testcase_gui). Add any missing shift dates and
            # persist the corrected case file so the solver receives the same view.
            try:
                calendar_data = self._ensure_shifts_in_calendar(calendar_data, shifts)
                case_data['calendar'] = calendar_data
                # Re-write saved input_case.json to reflect corrections
                try:
                    with open(case_file, 'w', encoding='utf-8') as f:
                        json.dump(case_data, f, indent=2)
                    logger.info(f"Rewrote input_case.json with calendar corrections: {case_file}")
                except Exception as _e:
                    logger.info(f"Failed to rewrite input_case.json after calendar corrections: {_e}")
            except Exception as e:
                logger.warning(f"Failed to ensure shifts in calendar: {e}")
            
            self._update_progress(run_id, 20, "Building optimization model...")
            
            # Build the OR-Tools model (simplified version)
            model_result = self._build_and_solve_model(
                constants, calendar_data, shifts, providers, run_config, run_id
            )
            
            self._update_progress(run_id, 90, "Generating output files...")
            
            # Save results
            result_file = run_output_dir / "results.json"
            with open(result_file, 'w') as f:
                json.dump(model_result, f, indent=2)
            
            # Generate Excel outputs (optional)
            try:
                # Coerce falsy/non-dict results to empty dict to avoid downstream errors
                if not isinstance(model_result, dict):
                    logger.warning(f"Model result is not a dict (type={type(model_result)}). Coercing to dict for output generation.")
                    model_result = model_result or {}

                self._generate_excel_outputs(model_result, run_output_dir)
            except Exception as e:
                logger.warning(f"Failed to generate Excel outputs: {e}")
            
            self._update_progress(run_id, 100, "Optimization completed successfully!")
            # Defensive: ensure model_result is a dict before manipulating debug_info
            if not isinstance(model_result, dict):
                logger.warning(f"Model result is not a dict (type={type(model_result)}). Coercing to dict.")
                # If model_result is falsy (None), replace with empty dict; if it's another type, coerce to dict wrapper
                model_result = model_result or {}
            # Ensure debug_info exists to signal which path was used
            model_result.setdefault('debug_info', {})
            model_result['debug_info'].setdefault('used_testcase_gui', False)

            return {
                "status": "success",
                "run_id": run_id,
                "output_directory": str(run_output_dir),
                "result": model_result
            }
            
        except Exception as e:
            logger.error(f"Solver error for run {run_id}: {str(e)}")
            logger.error(traceback.format_exc())
            
            self._update_progress(run_id, -1, f"Error: {str(e)}")
            
            return {
                "status": "error",
                "run_id": run_id,
                "message": str(e),
                "traceback": traceback.format_exc()
            }
    
    def _build_and_solve_model(self, constants: Dict, calendar: Dict, 
                             shifts: List, providers: List, run_config: Dict, run_id: str) -> Dict[str, Any]:
        """
        Core optimization path. Delegates solving to the original testcase_gui.py script
        and correctly handles both successful and error responses.
        """
        if not HAVE_ORIGINAL_SOLVER:
            raise RuntimeError("testcase_gui.py is not available. Cannot solve.")

        logger.info("Using original testcase_gui.py for solving...")
        
        run_config['out'] = run_config.get('out') or f"run_{run_id[:8]}"
        case = {
            "constants": constants or {},
            "calendar": calendar or {},
            "shifts": shifts or [],
            "providers": providers or [],
            "run": run_config or {}
        }
        
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json', encoding='utf-8') as tmp:
            json.dump(case, tmp)
            tmp_path = tmp.name
        
        try:
            self._update_progress(run_id, 30, "Calling testcase_gui.Solve_test_case...")

            # testcase_gui creates a relative 'out' directory. Ensure that
            # it's created under our solver_output folder by temporarily
            # changing cwd to self.output_dir before invoking the solver.
            current_cwd = os.getcwd()
            try:
                os.chdir(str(self.output_dir))
                solver_result = original_solver.Solve_test_case(os.path.abspath(tmp_path))
            finally:
                try:
                    os.chdir(current_cwd)
                except Exception:
                    pass
            
            if isinstance(solver_result, dict) and 'error' in solver_result:
                error_message = solver_result.get('error', 'Unknown solver script error')
                logger.error(f"testcase_gui.py returned an error: {error_message}")
                raise RuntimeError(f"Solver script failed: {error_message}")

            tables, meta = solver_result

            solutions = []
            for i, table_data in enumerate(tables):
                assignments = []
                for (s_idx, p_idx) in table_data.get('assignment', []):
                    shift = table_data['shifts'][s_idx]
                    provider = table_data['providers'][p_idx]
                    assignments.append({
                        "shift_id": shift['id'], "provider_name": provider['name'],
                        "date": shift['date'], "shift_type": shift.get('type', ''),
                        "start_time": shift.get('start', ''), "end_time": shift.get('end', '')
                    })
                
                objective = (meta.get('per_table', [])[i].get('objective') 
                             if i < len(meta.get('per_table', [])) else 0)

                solutions.append({ "assignments": assignments, "objective_value": objective })

            logger.info(f"✅ testcase_gui.py finished, found {len(solutions)} diverse solutions.")

            return { 'status': 'completed', 'solutions': solutions, 'solver_stats': meta.get('phase2', {}) }
        finally:
            os.remove(tmp_path)
    def _sanitize_calendar(self, calendar_obj: Dict[str, Any]) -> Dict[str, Any]:
        """
        Validate and sanitize calendar object. Ensures `days` is a list of ISO date
        strings and clamps day values that are out-of-range for their month to the
        month's last day. Returns a normalized calendar dict.

        Behavior:
        - If days is missing or not a list, returns calendar_obj unchanged.
        - For each day string, attempts to parse YYYY-MM-DD. If parsing fails,
          tries to recover by clamping numeric day to the month's max day.
        - Logs any corrections made and returns the cleaned calendar.
        """
        if not calendar_obj:
            return calendar_obj

        days = calendar_obj.get('days')
        if not isinstance(days, list):
            return calendar_obj

        cleaned: List[str] = []
        for idx, d in enumerate(days):
            if not isinstance(d, str):
                logger.warning(f"Non-string calendar entry at index {idx}: {d} - skipping")
                continue
            parts = d.split('-')
            if len(parts) != 3:
                logger.warning(f"Invalid ISO date format in calendar at index {idx}: {d} - skipping")
                continue
            try:
                y = int(parts[0])
                m = int(parts[1])
                dd = int(parts[2])
            except Exception:
                logger.warning(f"Non-numeric ISO parts in calendar at index {idx}: {d} - skipping")
                continue

            # Clamp month to 1..12
            if m < 1 or m > 12:
                logger.warning(f"Month out of range in calendar at index {idx}: {d} - skipping")
                continue

            # Determine last valid day for that month/year
            last_day = pycalendar.monthrange(y, m)[1]
            if dd < 1:
                logger.warning(f"Day out of range (<1) in calendar at index {idx}: {d} - skipping")
                continue
            if dd > last_day:
                corrected = f"{y:04d}-{m:02d}-{last_day:02d}"
                logger.warning(f"Correcting out-of-range calendar date at index {idx}: {d} -> {corrected}")
                cleaned.append(corrected)
            else:
                cleaned.append(f"{y:04d}-{m:02d}-{dd:02d}")

        # Return new calendar object with cleaned days and preserve other keys
        new_cal = dict(calendar_obj)
        new_cal['days'] = cleaned
        return new_cal

    def _ensure_shifts_in_calendar(self, calendar_obj: Dict[str, Any], shifts: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Ensure that every shift.date appears in calendar_obj['days']. If some
        shift dates are missing, log a warning and add them to the calendar days
        list. Returns the updated calendar object.
        """
        if not calendar_obj:
            calendar_obj = { 'days': [] }

        days = calendar_obj.get('days') or []
        if not isinstance(days, list):
            days = []

        existing = set(days)
        missing = set()
        for s in (shifts or []):
            d = s.get('date') if isinstance(s, dict) else None
            if isinstance(d, str) and d not in existing:
                missing.add(d)

        if not missing:
            return calendar_obj

        # Log and add missing dates
        missing_list = sorted(missing)
        logger.warning(f"Missing shift dates not present in calendar.days: {missing_list} - adding to calendar")

        # Merge and sort ISO date strings lexicographically (ISO order == chronological)
        merged = sorted(set(days) | missing)
        new_cal = dict(calendar_obj)
        new_cal['days'] = merged
        return new_cal

        # ---------- Built-in simplified OR-Tools model (fallback) ----------
        logger.info(f"Building CP-SAT model for run {run_id} (built-in)")
        
        # Initialize CP model
        model = cp_model.CpModel()
        
        # Extract configuration
        max_time = constants.get('solver', {}).get('max_time_in_seconds', 300)
        num_threads = constants.get('solver', {}).get('num_threads', 8)
        k_solutions = run_config.get('k', 5)
        
        self._update_progress(run_id, 30, f"Creating variables for {len(shifts)} shifts and {len(providers)} providers...")
        
        # Create decision variables
        shift_assignments = {}
        for shift in shifts:
            shift_id = shift['id']
            for provider in providers:
                provider_name = provider['name']
                # Binary variable: 1 if provider is assigned to shift, 0 otherwise
                var_name = f"assign_{provider_name}_{shift_id}"
                shift_assignments[(provider_name, shift_id)] = model.NewBoolVar(var_name)
        
        self._update_progress(run_id, 40, "Adding constraints...")
        
        # Constraint 1: Each shift must be assigned to exactly one provider
        for shift in shifts:
            shift_id = shift['id']
            model.Add(
                sum(shift_assignments[(provider['name'], shift_id)] for provider in providers) == 1
            )
        
        # Constraint 2: Provider availability and forbidden days
        shifts_by_date = collections.defaultdict(list)
        for shift in shifts:
            shifts_by_date[shift['date']].append(shift)
            
        for provider in providers:
            provider_name = provider['name']
            
            # Hard OFF days (forbidden)
            for off_day in provider.get('days_off', []):
                if off_day.get('type') == 'fixed':
                    date_str = off_day['date']
                    if date_str in shifts_by_date:
                        for shift in shifts_by_date[date_str]:
                            model.Add(shift_assignments[(provider_name, shift['id'])] == 0)
        
        # Constraint 3: At most one shift per provider per day
        for provider in providers:
            provider_name = provider['name']
            for date_str, day_shifts in shifts_by_date.items():
                if len(day_shifts) > 1:
                    model.Add(
                        sum(shift_assignments[(provider_name, shift['id'])] for shift in day_shifts) <= 1
                    )
        
        self._update_progress(run_id, 60, "Setting up objective function...")
        
        # Objective: Minimize violations and maximize preferences
        objective_terms = []
        
        # Soft constraints: Preferred days
        for provider in providers:
            provider_name = provider['name']
            for pref_day in provider.get('days_on', []):
                if pref_day.get('type') == 'prefer':
                    date_str = pref_day['date']
                    if date_str in shifts_by_date:
                        # Bonus for working on preferred days
                        for shift in shifts_by_date[date_str]:
                            objective_terms.append(
                                shift_assignments[(provider_name, shift['id'])] * 100
                            )
        
        # Fairness: Try to balance workload
        provider_workloads = []
        for provider in providers:
            provider_name = provider['name']
            workload = sum(
                shift_assignments[(provider_name, shift['id'])] 
                for shift in shifts
            )
            provider_workloads.append(workload)
        
        # Add workload balancing terms (CP-SAT-safe)
        if provider_workloads:
            # Create an integer variable for the total workload and constrain it
            # to equal the (symbolic) sum of provider workloads.
            total_workload = model.NewIntVar(0, len(shifts) * len(providers), 'total_workload')
            model.Add(total_workload == sum(provider_workloads))

            # Represent the average as an IntVar and relate it to total_workload
            # via multiplication by the number of providers. This avoids doing
            # Python-side integer division on OR-Tools expressions.
            n_providers = len(provider_workloads)
            avg_workload = model.NewIntVar(0, len(shifts), 'avg_workload')
            remainder = model.NewIntVar(0, max(0, n_providers - 1), 'avg_remainder')
            model.Add(total_workload == avg_workload * n_providers + remainder)

            for idx, workload in enumerate(provider_workloads):
                # Use unique var names per provider to avoid accidental reuse
                suffix = f"_{idx}"
                deviation = model.NewIntVar(-len(shifts), len(shifts), f'deviation{suffix}')
                model.Add(deviation == workload - avg_workload)
                abs_deviation = model.NewIntVar(0, len(shifts), f'abs_deviation{suffix}')
                model.AddAbsEquality(abs_deviation, deviation)
                # Minimize absolute deviation from average
                objective_terms.append(-abs_deviation)
        
        if objective_terms:
            model.Maximize(sum(objective_terms))
        
        self._update_progress(run_id, 70, "Solving optimization model...")
        
        # Solve the model
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = max_time
        solver.parameters.num_search_workers = num_threads
        
        # Collect multiple solutions if requested
        if k_solutions > 1:
            solution_collector = SolutionCollector(shift_assignments, shifts, providers, k_solutions)
            status = solver.SolveWithSolutionCallback(model, solution_collector)
            solutions = solution_collector.get_solutions()
        else:
            status = solver.Solve(model)
            solutions = []
            
            if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
                # Extract single solution
                assignments = []
                for shift in shifts:
                    shift_id = shift['id']
                    for provider in providers:
                        provider_name = provider['name']
                        if solver.Value(shift_assignments[(provider_name, shift_id)]):
                            assignments.append({
                                "shift_id": shift_id,
                                "provider_name": provider_name,
                                "date": shift['date'],
                                "shift_type": shift.get('type', ''),
                                "start_time": shift.get('start', ''),
                                "end_time": shift.get('end', '')
                            })
                
                solutions.append({
                    "assignments": assignments,
                    "objective_value": solver.ObjectiveValue() if solver.ObjectiveValue() else 0
                })
        
        self._update_progress(run_id, 85, "Processing results...")
        
        # Generate statistics
        total_assignments = sum(len(sol.get('assignments', [])) for sol in solutions)
        provider_stats = collections.Counter()
        shift_type_stats = collections.Counter()
        
        for solution in solutions:
            for assignment in solution.get('assignments', []):
                provider_stats[assignment['provider_name']] += 1
                shift_type_stats[assignment['shift_type']] += 1
        
        result = {
            "solver_status": self._get_status_name(status),
            "solutions_found": len(solutions),
            "solutions": solutions,
            "statistics": {
                "total_shifts": len(shifts),
                "total_providers": len(providers),
                "total_assignments": total_assignments,
                "provider_workload": dict(provider_stats),
                "shift_type_coverage": dict(shift_type_stats),
                "runtime_seconds": solver.WallTime(),
                "objective_value": solver.ObjectiveValue() if solutions else 0
            },
            "solver_info": {
                "status": self._get_status_name(status),
                "runtime": f"{solver.WallTime():.2f} seconds",
                "num_conflicts": solver.NumConflicts(),
                "num_branches": solver.NumBranches()
            }
        }

        # Provide approximate variable/constraint counts useful for tests
        try:
            # Variables ~ assignment vars + 2 per provider (balancing vars)
            var_count = len(shift_assignments) + max(0, len(providers) * 2)
            cons_count = len(shifts)  # one per shift exactly-one
            cons_count += sum(1 for _ in providers) * len(shifts_by_date)  # at-most-one per provider/day
            result.setdefault("solver_info", {})
            result["solver_info"].update({
                "variables": var_count,
                "constraints": cons_count
            })
        except Exception:
            pass
        
        return result

    def _to_webapp_response(self, model_result: Dict[str, Any], run_id: str) -> Dict[str, Any]:
        """Normalize model_result into response expected by existing local solver clients."""
        # Extract solutions
        solutions: List[Dict[str, Any]] = []
        if isinstance(model_result, dict):
            if isinstance(model_result.get('solutions'), list):
                solutions = model_result['solutions']
            elif isinstance(model_result.get('results'), dict) and isinstance(model_result['results'].get('solutions'), list):
                solutions = model_result['results']['solutions']

        # Build solver_stats if absent
        solver_stats = model_result.get('solver_stats') or {}
        if not solver_stats:
            exec_ms = 0
            try:
                rt = model_result.get('statistics', {}).get('runtime_seconds')
                if rt is not None:
                    exec_ms = int(float(rt) * 1000)
            except Exception:
                pass
            impl = model_result.get('solver_info', {}).get('implementation') if isinstance(model_result.get('solver_info'), dict) else None
            solver_type = 'real_scheduler_sat_core' if impl == 'testcase_gui.py' else 'ortools_fastapi'
            solver_stats = {
                'total_solutions': len(solutions),
                'execution_time_ms': exec_ms,
                'solver_type': solver_type,
                'status': model_result.get('solver_status', 'UNKNOWN')
            }

        payload = {
            'status': 'completed',
            'message': 'Optimization completed',
            'run_id': run_id,
            'progress': 100,
            'results': {
                'solutions': solutions,
                'solver_stats': solver_stats
            },
            'statistics': model_result.get('statistics', {})
        }
        if 'solver_info' in model_result:
            payload['solver_info'] = model_result['solver_info']
        return payload

    def _coerce_tcg_result(self, tcg_out: Any) -> Optional[Dict[str, Any]]:
        """Best-effort adapter to transform testcase_gui outputs into API result schema.
        Accepts either a dict with solutions/assignments or a custom structure.
        Returns normalized dict or None if unable to adapt.
        """
        try:
            # If already in expected shape
            if isinstance(tcg_out, dict) and (
                'solutions' in tcg_out or (
                    'results' in tcg_out and isinstance(tcg_out['results'], dict) and 'solutions' in tcg_out['results']
                )
            ):
                return tcg_out

            # Common pattern: list of solutions or a pool with assignments
            if isinstance(tcg_out, list):
                sols = []
                for idx, sol in enumerate(tcg_out):
                    if isinstance(sol, dict) and 'assignments' in sol:
                        sols.append({
                            'assignments': sol['assignments'],
                            'objective_value': sol.get('objective_value', 0),
                        })
                if sols:
                    return {
                        'solver_status': 'UNKNOWN',
                        'solutions_found': len(sols),
                        'solutions': sols,
                        'statistics': {}
                    }

            # Sometimes a tuple like (solutions, stats)
            if isinstance(tcg_out, tuple) and tcg_out:
                primary = tcg_out[0]
                return self._coerce_tcg_result(primary)

        except Exception as e:
            logger.debug(f"Coercion of testcase_gui output failed: {e}")
        return None
    
    def _get_status_name(self, status) -> str:
        """Convert CP solver status to readable string"""
        status_names = {
            cp_model.OPTIMAL: "OPTIMAL",
            cp_model.FEASIBLE: "FEASIBLE", 
            cp_model.INFEASIBLE: "INFEASIBLE",
            cp_model.UNKNOWN: "UNKNOWN",
            cp_model.MODEL_INVALID: "MODEL_INVALID"
        }
        return status_names.get(status, f"UNKNOWN_STATUS_{status}")
    
    def _generate_excel_outputs(self, result: Dict[str, Any], output_dir: Path):
        """Generate Excel files for the solution (optional feature)"""
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Schedule"
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Headers
            ws.append(["Date", "Shift Type", "Provider", "Start Time", "End Time"])
            
            # Add assignments from first solution (if available)
            sols = result.get('solutions') if isinstance(result, dict) else None
            if not sols:
                logger.info("No solutions present in result; creating empty Excel file.")
                ws = wb.create_sheet("No_Solution_Found")
                ws.append(["Status", result.get('solver_stats', {}).get('final_status', 'NO_SOLUTION')])
            else:
                for idx, solution in enumerate(sols):
                    ws = wb.create_sheet(f"Schedule_{idx + 1}")
                    ws.append(["Date", "Shift Type", "Provider", "Start Time", "End Time", "Shift ID"])
                    
                    assignments = solution.get('assignments', [])
                    # Sort assignments by date and time for readability
                    assignments.sort(key=lambda x: (x.get('date', ''), x.get('start_time', '')))
                    
                    for assignment in assignments:
                        ws.append([
                            assignment.get('date', ''),
                            assignment.get('shift_type', ''),
                            assignment.get('provider_name', ''),
                            assignment.get('start_time', '').split('T')[-1], # Show only HH:MM:SS
                            assignment.get('end_time', '').split('T')[-1],   # Show only HH:MM:SS
                            assignment.get('shift_id', '')
                        ])
            
            
            excel_file = output_dir / "schedule.xlsx"
            wb.save(excel_file)
            logger.info(f"Generated Excel output with {len(sols) if sols else 0} solution(s): {excel_file}")
            
        except Exception as e:
            logger.warning(f"Failed to generate Excel: {e}")
            logger.error(traceback.format_exc())
    
    def _update_progress(self, run_id: str, progress: float, message: str):
        """Update progress and notify WebSocket clients"""
        if run_id in active_runs:
            active_runs[run_id]['progress'] = progress
            active_runs[run_id]['message'] = message
            active_runs[run_id]['updated_at'] = datetime.now().isoformat()
        
        # Notify WebSocket clients
        if run_id in websocket_connections:
            try:
                asyncio.create_task(self._send_progress_update(run_id, progress, message))
            except Exception as e:
                logger.warning(f"Failed to send WebSocket update: {e}")
        
        logger.info(f"Run {run_id}: {progress}% - {message}")
    
    async def _send_progress_update(self, run_id: str, progress: float, message: str):
        """Send progress update via WebSocket"""
        if run_id in websocket_connections:
            try:
                await websocket_connections[run_id].send_text(json.dumps({
                    "type": "progress",
                    "run_id": run_id,
                    "progress": progress,
                    "message": message,
                    "timestamp": datetime.now().isoformat()
                }))
            except Exception as e:
                logger.warning(f"WebSocket send failed: {e}")

class SolutionCollector(cp_model.CpSolverSolutionCallback):
    """Collect multiple solutions from the CP solver"""
    
    def __init__(self, shift_assignments: Dict, shifts: List, providers: List, max_solutions: int):
        cp_model.CpSolverSolutionCallback.__init__(self)
        self._shift_assignments = shift_assignments
        self._shifts = shifts
        self._providers = providers
        self._solutions = []
        self._max_solutions = max_solutions
    
    def on_solution_callback(self):
        if len(self._solutions) >= self._max_solutions:
            self.StopSearch()
            return
        
        # Extract current solution
        assignments = []
        for shift in self._shifts:
            shift_id = shift['id']
            for provider in self._providers:
                provider_name = provider['name']
                if self.Value(self._shift_assignments[(provider_name, shift_id)]):
                    assignments.append({
                        "shift_id": shift_id,
                        "provider_name": provider_name,
                        "date": shift['date'],
                        "shift_type": shift.get('type', ''),
                        "start_time": shift.get('start', ''),
                        "end_time": shift.get('end', '')
                    })
        
        self._solutions.append({
            "assignments": assignments,
            "objective_value": self.ObjectiveValue()
        })
    
    def get_solutions(self) -> List[Dict[str, Any]]:
        return self._solutions

# Initialize solver
solver = AdvancedSchedulingSolver()

from fastapi import BackgroundTasks
# REST API Endpoints
@app.post("/solve")
async def solve_schedule(case: SchedulingCase, background_tasks: BackgroundTasks):
    """Submit a scheduling case for optimization (synchronous)."""
    try:
        run_id = str(uuid.uuid4())
        case_dict = case.dict()

        active_runs[run_id] = {
            "status": "running",
            "progress": 0,
            "message": "Optimization started",
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }

        # Run optimization synchronously for local usage
        result = await solver.solve_async(case_dict, run_id)
        active_runs[run_id].update({
            "status": result.get("status", "success"),
            "progress": 100 if result.get("status") == "success" else -1,
            "message": "Completed" if result.get("status") == "success" else result.get("message", "Failed"),
            "result": result,
            "completed_at": datetime.now().isoformat()
        })
        background_tasks.add_task(run_optimization, case_dict, run_id)
        # Normalize to the shape expected by the web app/tests
        model_result = result.get("result", {})
        return JSONResponse(
            status_code=202, # HTTP 202 Accepted
            content={
                "status": "accepted",
                "message": "Optimization started in the background.",
                "run_id": run_id,
                "solver_service_url": f"http://localhost:8000/status/{run_id}",
                "websocket_url": f"ws://localhost:8000/ws/{run_id}",
            }
        )
    except Exception as e:
        logger.error(f"API error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
async def _send_final_result(run_id: str, result: Dict[str, Any]):
    """Send the final result payload via WebSocket."""
    if run_id in websocket_connections:
        try:
            # The frontend expects a specific structure, so we normalize it here.
            normalized_payload = solver._to_webapp_response(result.get("result", {}), run_id)
            
            await websocket_connections[run_id].send_text(json.dumps({
                "type": "result",
                "run_id": run_id,
                "payload": normalized_payload,
                "timestamp": datetime.now().isoformat()
            }))
            logger.info(f"Sent final result to WebSocket for run {run_id}")
        except Exception as e:
            logger.warning(f"Failed to send final result to WebSocket for run {run_id}: {e}")

async def run_optimization(case_data: Dict[str, Any], run_id: str):
    """Background task for running optimization"""
    try:
        active_runs[run_id]["status"] = "running"
        result = await solver.solve_async(case_data, run_id)
        
        # Update final status
        active_runs[run_id].update({
            "status": result["status"],
            "progress": 100 if result["status"] == "success" else -1,
            "message": "Completed" if result["status"] == "success" else result.get("message", "Failed"),
            "result": result,
            "completed_at": datetime.now().isoformat()
        })
        if result["status"] == "success":
            await _send_final_result(run_id, result)
        
    except Exception as e:
        logger.error(f"Background optimization failed: {e}")
        active_runs[run_id].update({
            "status": "error",
            "progress": -1,
            "message": str(e),
            "completed_at": datetime.now().isoformat()
        })

@app.get("/status/{run_id}", response_model=SolverStatus)
async def get_status(run_id: str):
    """Get status of a specific optimization run"""
    if run_id not in active_runs:
        raise HTTPException(status_code=404, detail="Run not found")
    
    run_data = active_runs[run_id]
    return SolverStatus(
        status=run_data["status"],
        message=run_data["message"],
        run_id=run_id,
        progress=run_data.get("progress", 0),
        results=run_data.get("result")
    )

@app.get("/runs")
async def list_runs():
    """List all optimization runs"""
    return {
        "runs": [
            {
                "run_id": run_id,
                "status": data["status"],
                "progress": data.get("progress", 0),
                "message": data["message"],
                "created_at": data["created_at"],
                "updated_at": data["updated_at"]
            }
            for run_id, data in active_runs.items()
        ]
    }

@app.websocket("/ws/{run_id}")
async def websocket_endpoint(websocket: WebSocket, run_id: str):
    """WebSocket for real-time progress updates"""
    await websocket.accept()
    websocket_connections[run_id] = websocket
    
    try:
        # Send initial status if run exists
        if run_id in active_runs:
            run_data = active_runs[run_id]
            await websocket.send_text(json.dumps({
                "type": "status",
                "run_id": run_id,
                "status": run_data["status"],
                "progress": run_data.get("progress", 0),
                "message": run_data["message"]
            }))
        
        # Keep connection alive and listen for client messages
        while True:
            data = await websocket.receive_text()
            # Echo back for keep-alive
            await websocket.send_text(json.dumps({
                "type": "ping",
                "message": "Connection alive"
            }))
            
    except WebSocketDisconnect:
        logger.info(f"WebSocket disconnected for run {run_id}")
    finally:
        if run_id in websocket_connections:
            del websocket_connections[run_id]

@app.get("/output/{run_id}")
async def get_output_files(run_id: str):
    """Get list of output files for a specific run"""
    if run_id not in active_runs:
        raise HTTPException(status_code=404, detail="Run not found")
    
    run_dir = solver.output_dir / run_id
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail="Output directory not found")
    
    files = list(run_dir.glob("*"))
    return {
        "run_id": run_id,
        "output_directory": str(run_dir),
        "files": [
            {
                "name": f.name,
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
            }
            for f in files if f.is_file()
        ]
    }

@app.get("/download/{run_id}/{filename}")
async def download_file(run_id: str, filename: str):
    """Download a specific output file"""
    if run_id not in active_runs:
        raise HTTPException(status_code=404, detail="Run not found")
    
    file_path = solver.output_dir / run_id / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(path=file_path, filename=filename)

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "ok",
        "message": "FastAPI Scheduling Solver Service is running",
        "timestamp": datetime.now().isoformat(),
        "active_runs": len(active_runs),
        "websocket_connections": len(websocket_connections)
    }

@app.get("/")
async def root():
    """API information"""
    return {
        "title": "Medical Staff Scheduling Solver API",
        "version": "2.0.0",
        "description": "High-performance optimization service using FastAPI and OR-Tools",
        "endpoints": {
            "POST /solve": "Submit scheduling case for optimization",
            "GET /status/{run_id}": "Get optimization status",
            "GET /runs": "List all runs",
            "WebSocket /ws/{run_id}": "Real-time progress updates",
            "GET /output/{run_id}": "List output files",
            "GET /download/{run_id}/{filename}": "Download output file",
            "GET /health": "Health check"
        },
        "documentation": "/docs"
    }
    
@app.get("/results/folders")
async def list_result_folders():
    """List all result folders in the solver_output directory."""
    output_dir = Path("solver_output")
    if not output_dir.exists():
        return {"folders": []}
    
    folders = []
    for d in output_dir.iterdir():
        if d.is_dir():
            try:
                stat = d.stat()
                file_count = len(list(d.glob('*')))
                folders.append({
                    "name": d.name,
                    "path": str(d),
                    "created": stat.st_ctime,
                    "fileCount": file_count
                })
            except Exception as e:
                logger.warning(f"Could not stat folder {d}: {e}")
    
    # Sort by creation time, newest first
    folders.sort(key=lambda x: x["created"], reverse=True)
    return {"folders": folders}

if __name__ == "__main__":
    print("🚀 Starting Medical Staff Scheduling Solver Service (FastAPI)")
    print("📊 Service URL: http://localhost:8000")
    print("📚 API Documentation: http://localhost:8000/docs")
    print("🔌 WebSocket: ws://localhost:8000/ws/{run_id}")
    print("\n🎯 Endpoints:")
    print("  POST /solve              - Submit optimization case")
    print("  GET  /status/{run_id}    - Get run status") 
    print("  GET  /runs               - List all runs")
    print("  WebSocket /ws/{run_id}   - Real-time updates")
    print("  GET  /health             - Health check")
    print("\n⚡ Press Ctrl+C to stop the service")
    
    uvicorn.run(
        "fastapi_solver_service:app",
        host="0.0.0.0",
        port=8000,
        reload=False,  # Disable in production
        log_level="info"
    )