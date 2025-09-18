#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scheduler Testcase Builder GUI
- Merged format: { constants, run, calendar, shifts, providers }
- Tabs:
    1) Run (output dir, run controls, live log)
    2) Calendar (weekend days, month generator, calendar day list)
    3) Shifts (per-day shift table, add/update/delete; add across ALL days)
    4) Providers (edit provider, OFF dates (fixed/prefer), ON dates (fixed/prefer) + shift-type picker, prefs summary)
    5) Config (solver constants / weights / objective editors)
- Fixed: naming collision between Calendar and Providers day listboxes.
  * Calendar uses self.lst_days_cal
  * Providers now uses two panes: self.lst_days_off and self.lst_days_on
- Clear prefs on selected days works and updates summary/log.
- Color-coded OFF/ON day lists with legends.
"""
from __future__ import annotations

import json
import os
import sys
import threading
import subprocess
import queue
import random
from pathlib import Path
from datetime import date, timedelta, datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

import argparse, json, os, re, sys, subprocess, traceback
import datetime as dt
from collections import defaultdict
from typing import Dict, Any, List

import logging
from logging import Logger

from ortools.sat.python import cp_model
from openpyxl import Workbook
CHOSPITAL = ""
SCALE = 1
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
diagnose_schedule_api.py  — colorized, expanded diagnostics (multi-sheet aware)

Programmatic usage (no CLI):
  from diagnose_schedule_api import run
  run(case_path, schedule_path, no_color=False, preview=8)

Behavior is identical to the CLI version you shared:
- If the schedule is an Excel workbook (.xlsx/.xlsm), ALL suitable sheets
  (those having ShiftID and Provider/Assignee columns) are diagnosed.
  One report file is written per sheet next to the schedule file:
     <schedule_stem>__<SheetName>.diagnose.txt
- CSV/JSON behave as before (single report). For single-item inputs, prints to stdout.

Deps:
  - colorama (optional for terminal colors)
  - openpyxl (only if reading .xlsx/.xlsm)
"""

import json, csv, sys, collections, io
from pathlib import Path
from datetime import date, datetime

# -------------- Colors --------------
_USE_COLOR = True
try:
    from colorama import init as _colorama_init, Fore, Style
    _colorama_init()
except Exception:
    class _Dummy:
        def __getattr__(self, k): return ""
    Fore = Style = _Dummy()

def _c_ok(s):   return (Fore.GREEN + s + Style.RESET_ALL) if _USE_COLOR else s
def _c_fail(s): return (Fore.RED + s + Style.RESET_ALL)   if _USE_COLOR else s
def _c_warn(s): return (Fore.YELLOW + s + Style.RESET_ALL)if _USE_COLOR else s
def _c_head(s): return (Fore.CYAN + s + Style.RESET_ALL)  if _USE_COLOR else s
def _c_dim(s):  return (Style.DIM + s + Style.RESET_ALL)  if _USE_COLOR else s

# -------------- Robust readers --------------

_PREFERRED_ENCODINGS = ("utf-8", "utf-8-sig", "cp1252", "latin-1")

def _read_text_best_effort(path: str):
    for enc in _PREFERRED_ENCODINGS:
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                return f.read(), enc, False
        except UnicodeDecodeError:
            continue
    data = Path(path).read_bytes()
    txt = data.decode("latin-1", errors="replace")
    return txt, "latin-1-replace", True

def _read_json_best_effort(path: str):
    last_err = None
    for enc in _PREFERRED_ENCODINGS:
        try:
            with open(path, "r", encoding=enc) as f:
                return json.load(f), enc
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except json.JSONDecodeError:
            raise
    raw = Path(path).read_bytes().decode("latin-1", errors="replace")
    return json.loads(raw), "latin-1-replace"

# -------------- Case loader --------------

def load_case(path):
    case, enc = _read_json_best_effort(path)
    cal = case.get("calendar", {}) or {}
    days = list(cal.get("days", []))
    weekend_days = cal.get("weekend_days", ["Saturday","Sunday"])
    shifts = case.get("shifts", []) or []
    providers = case.get("providers", []) or []
    for p in providers:
        p.setdefault("name", p.get("id", ""))
        p.setdefault("type", "MD")
        p.setdefault("forbidden_days_hard", [])
        p.setdefault("forbidden_days_soft", [])
        p.setdefault("preferred_days_hard", {})
        p.setdefault("preferred_days_soft", {})
        lim = p.setdefault("limits", {})
        lim.setdefault("min_total", 0)
        lim.setdefault("max_total", None)
        lim.setdefault("type_ranges", {})
        if not isinstance(p.get("max_consecutive_days", None), int):
            p["max_consecutive_days"] = None
        if isinstance(p["preferred_days_hard"], list):
            p["preferred_days_hard"] = {}
        if isinstance(p["preferred_days_soft"], list):
            p["preferred_days_soft"] = {}
    return {
        "calendar_days": days,
        "weekend_days": weekend_days,
        "shifts": shifts,
        "providers": providers,
        "case_encoding": enc
    }

# -------------- Schedule loaders --------------

def infer_schedule_from_json(obj):
    if isinstance(obj, dict) and "assignments" in obj and isinstance(obj["assignments"], list):
        out = {}
        for row in obj["assignments"]:
            sid = row.get("shift_id") or row.get("id") or row.get("shift")
            prov = row.get("provider") or row.get("provider_name") or row.get("name")
            if sid and prov:
                out.setdefault(str(sid).strip(), []).append(str(prov).strip())
        return out
    if isinstance(obj, dict):
        ok = True
        for k in obj.keys():
            if not isinstance(k, str):
                ok = False; break
        if ok:
            out = {}
            for k, v in obj.items():
                if isinstance(v, list):
                    out[str(k).strip()] = [str(x).strip() for x in v if str(x).strip()]
                elif isinstance(v, str):
                    out[str(k).strip()] = [v.strip()]
            if out:
                return out
    return None

def _find_candidate_columns(header_lower):
    def find_col(names):
        for i, h in enumerate(header_lower):
            if h in names:
                return i
        return None
    c_shift = find_col({"shiftid", "shift id", "id", "shift"})
    c_prov  = find_col({"assignee", "provider", "provider_name", "name"})
    return c_shift, c_prov

def _load_schedules_from_xlsx(path: str):
    """Return list of (sheet_name, schedule_map). One entry per suitable sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to read .xlsx/.xlsm files. Install via: pip install openpyxl")

    wb = load_workbook(path, data_only=True, read_only=True)
    results = []

    for ws in wb.worksheets:
        # sniff header
        header = None
        for row in ws.iter_rows(max_row=10, values_only=True):
            if not row: continue
            nonempty = [c for c in row if c not in (None, "")]
            if len(nonempty) < 2: continue
            header = [str(c).strip() if c is not None else "" for c in row]
            break
        if not header:
            continue
        lower = [h.lower() for h in header]
        c_shift, c_prov = _find_candidate_columns(lower)
        if c_shift is None or c_prov is None:
            continue

        # read rows
        first = True
        sheet_map = {}
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row:
                continue
            sid = row[c_shift] if c_shift < len(row) else None
            prov = row[c_prov] if c_prov < len(row) else None
            if sid is None or str(sid).strip() == "":
                continue
            if prov is None or str(prov).strip() == "" or str(prov).strip().upper() == "UNFILLED":
                continue
            sid = str(sid).strip()
            prov = str(prov).strip()
            sheet_map.setdefault(sid, []).append(prov)

        if sheet_map:
            results.append((ws.title, sheet_map))

    if not results:
        sheet_names = ", ".join(ws.title for ws in wb.worksheets)
        raise ValueError(
            f"No suitable worksheets in {path}. Need headers: ShiftID & Provider/Assignee. "
            f"Sheets present: [{sheet_names}]"
        )
    return results

def _load_schedule_from_csv(path: str):
    text, enc, had_rep = _read_text_best_effort(str(path))
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        class _D(csv.excel):
            delimiter = ','
        dialect = _D()
    try:
        has_header = csv.Sniffer().has_header(sample)
    except Exception:
        has_header = True
    fobj = io.StringIO(text)
    reader = csv.DictReader(fobj, dialect=dialect) if has_header else csv.DictReader(fobj, fieldnames=None, dialect=dialect)
    col_shift = col_provider = None
    if reader.fieldnames:
        lower = [c.lower() for c in reader.fieldnames]
        c_shift, c_prov = _find_candidate_columns(lower)
        if c_shift is not None: col_shift = reader.fieldnames[c_shift]
        if c_prov  is not None: col_provider = reader.fieldnames[c_prov]
    if not col_shift or not col_provider:
        raise ValueError(f"CSV must include columns for shift and provider; got fields: {reader.fieldnames} (encoding={enc})")
    out = {}
    for row in reader:
        sid = (row.get(col_shift) or "").strip()
        prov = (row.get(col_provider) or "").strip()
        if sid and prov and prov.upper() != "UNFILLED":
            out.setdefault(sid, []).append(prov)
    return out, enc

def load_schedules(path_str: str):
    """
    Returns (items, source_descr)
    items is a list of (label, schedule_map).
      - For xlsx/xlsm: one per suitable sheet.
      - For json/csv: single item [(stem, schedule_map)].
    """
    p = Path(path_str)
    ext = p.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        sheets = _load_schedules_from_xlsx(str(p))
        return sheets, "xlsx"
    if ext == ".json":
        obj, enc = _read_json_best_effort(str(p))
        m = infer_schedule_from_json(obj)
        if m is None:
            raise ValueError("Unrecognized JSON schedule format. Expected {'assignments': [...]} or {shift_id: provider}.")
        return [(p.stem, m)], f"json({enc})"
    # CSV (or tsv-like)
    m, enc = _load_schedule_from_csv(str(p))
    return [(p.stem, m)], f"csv({enc})"

# -------------- Helpers --------------

def iso_weekday_name(dstr):
    y,m,dd = map(int, dstr.split("-"))
    return date(y,m,dd).strftime("%A")

def _to_date(dstr: str) -> date:
    y,m,dd = map(int, dstr.split("-"))
    return date(y,m,dd)

def _cluster_sizes(worked_days):
    if not worked_days:
        return []
    dd = sorted(_to_date(d) for d in set(worked_days))
    sizes = []; run = 1
    for i in range(1, len(dd)):
        if (dd[i] - dd[i-1]).days == 1:
            run += 1
        else:
            sizes.append(run); run = 1
    sizes.append(run)
    return sizes

# -------------- Diagnosis --------------

def diagnose(case, schedule_map, stream=sys.stdout, preview_limit=8, banner=None):
    days = case["calendar_days"]
    shifts = case["shifts"]
    providers = case["providers"]
    weekend_names = set(case["weekend_days"] or ["Saturday","Sunday"])

    if banner:
        print(_c_head(banner), file=stream)

    # Indexes
    shifts_by_id = {sh["id"]: sh for sh in shifts if "id" in sh}
    providers_by_name = {p["name"]: p for p in providers if "name" in p}
    day_to_shifts = collections.defaultdict(list)
    for s in shifts:
        day_to_shifts[s["date"]].append(s["id"])

    shift_date = {sid: sh["date"] for sid, sh in shifts_by_id.items()}
    # FIXED: correct dict comprehension (remove bad/duplicate line)
    shift_type = {sid: sh.get("type","") for sid, sh in shifts_by_id.items()}
    shift_allowed_types = {sid: set(sh.get("allowed_provider_types", ["MD"])) for sid, sh in shifts_by_id.items()}

    prov_day_to_shifts = collections.defaultdict(lambda: collections.defaultdict(list))
    for sid, provs in schedule_map.items():
        if sid not in shifts_by_id:
            continue
        d = shift_date[sid]
        for prov in provs:
            prov_day_to_shifts[prov][d].append(sid)

    checks = []
    def add_check(name, ok, details=""):
        checks.append((name, bool(ok), details))

    # 1) All shifts filled exactly once
    unfilled, overfilled = [], []
    for sid in shifts_by_id:
        n = len(schedule_map.get(sid, []))
        if n == 0:
            unfilled.append(sid)
        elif n > 1:
            overfilled.append((sid, n))
    add_check("All shifts filled exactly once", len(unfilled)==0 and len(overfilled)==0,
              (f"Unfilled:{len(unfilled)}; " if unfilled else "") + (f"Overfilled:{len(overfilled)}" if overfilled else ""))

    # 2) Unknown shift IDs
    unknown_shift_ids = [sid for sid in schedule_map if sid not in shifts_by_id]
    add_check("No unknown shift IDs in schedule", len(unknown_shift_ids)==0, f"Unknown IDs: {len(unknown_shift_ids)}")

    # 3) Provider exists
    unknown_providers = [prov for prov in prov_day_to_shifts.keys() if prov not in providers_by_name]
    add_check("All providers exist", len(unknown_providers)==0,
              f"Unknown providers: {', '.join(unknown_providers[:5])}{'...' if len(unknown_providers)>5 else ''}")

    # 4) Provider type allowed by shift
    bad_allowed = []
    for prov, by_day in prov_day_to_shifts.items():
        p = providers_by_name.get(prov)
        if not p: continue
        ptype = p.get("type","MD")
    for d, sids in by_day.items():
            for sid in sids:
                allowed = shift_allowed_types.get(sid) or set()
                if allowed and ptype not in allowed:
                    bad_allowed.append((prov, sid, ptype, sorted(allowed)))
    add_check("Provider type allowed for each assigned shift", len(bad_allowed)==0, f"Violations: {len(bad_allowed)}")

    # 5) Forbidden (hard-off) days
    bad_forbidden = []
    for prov, by_day in prov_day_to_shifts.items():
        p = providers_by_name.get(prov)
        if not p: continue
        forb = set(p.get("forbidden_days_hard", []))
        for d in by_day:
            if d in forb:
                bad_forbidden.append((prov, d, by_day[d]))
    add_check("Providers NOT scheduled on forbidden (hard-off) days", len(bad_forbidden)==0, f"Violations: {len(bad_forbidden)}")

    # 6) At most one shift per provider per day
    multi_same_day = []
    for prov, by_day in prov_day_to_shifts.items():
        for d, sids in by_day.items():
            if len(sids) > 1:
                multi_same_day.append((prov, d, sids))
    add_check("At most one shift per provider per day", len(multi_same_day)==0, f"Violations: {len(multi_same_day)}")

    # 7) Max consecutive days
    bad_consec = []
    for prov, by_day in prov_day_to_shifts.items():
        p = providers_by_name.get(prov)
        if not p: continue
        K = p.get("max_consecutive_days", None)
        if not (isinstance(K, int) and K > 0):
            continue
        worked_dates = sorted(map(lambda s: _to_date(s), by_day.keys()))
        longest = 0; cur = 0; prev = None
        for dt in worked_dates:
            if prev is None or (dt - prev).days == 1:
                cur += 1
            else:
                longest = max(longest, cur)
                cur = 1
            prev = dt
        longest = max(longest, cur)
        if longest > K:
            bad_consec.append((prov, longest, K))
    add_check("Max consecutive working days respected", len(bad_consec)==0, f"Violations: {len(bad_consec)}")

    # 8) Preferred-days HARD respected when working
    shift_type_map = {sid: sh.get("type","") for sid, sh in shifts_by_id.items()}
    bad_pref_hard = []
    for prov, by_day in prov_day_to_shifts.items():
        p = providers_by_name.get(prov)
        if not p: continue
        pref_map = p.get("preferred_days_hard", {}) or {}
        for d, sids in by_day.items():
            prefs = set(pref_map.get(d, []))
            if not prefs:
                continue
            for sid in sids:
                t = shift_type_map.get(sid, "")
                if t not in prefs:
                    bad_pref_hard.append((prov, d, sid, t, sorted(prefs)))
    add_check("Preferred-days HARD respected when working", len(bad_pref_hard)==0, f"Violations: {len(bad_pref_hard)}")

    # 9) Required-days HARD satisfied
    hard_on_misses = []
    for prov, p in providers_by_name.items():
        pref_map = p.get("preferred_days_hard", {}) or {}
        by_day = prov_day_to_shifts.get(prov, {})
        for d, req_types in pref_map.items():
            if not req_types:
                continue
            assigned_sids = by_day.get(d, [])
            assigned_types = [shift_type_map.get(sid, "") for sid in assigned_sids]
            if not assigned_sids or not any(t in set(req_types) for t in assigned_types):
                hard_on_misses.append((prov, d, sorted(set(req_types)), sorted(assigned_types)))
    add_check("Required-days HARD satisfied (worked one of required types)", len(hard_on_misses)==0,
              f"Violations: {len(hard_on_misses)}")

    # 10) Min/Max total shifts per provider (hard)
    minmax_viol = []
    prov_totals = {prov: sum(len(v) for v in by_day.values()) for prov, by_day in prov_day_to_shifts.items()}
    for prov in providers_by_name:
        total = prov_totals.get(prov, 0)
        lim = providers_by_name[prov].get("limits", {}) or {}
        mn = lim.get("min_total", 0)
        mx = lim.get("max_total", None)
        if mx is not None and total > mx:
            minmax_viol.append((prov, total, mn, mx))
        elif total < (mn or 0):
            minmax_viol.append((prov, total, mn, mx))
    add_check("Provider min_total/max_total respected", len(minmax_viol)==0, f"Violations: {len(minmax_viol)}")

    # 11) Per-type min/max ranges (hard)
    type_range_viol = []
    prov_type_counts = {}
    for prov, by_day in prov_day_to_shifts.items():
        counter = collections.Counter()
        for d, sids in by_day.items():
            for sid in sids:
                counter[shift_type_map.get(sid, "")] += 1
        prov_type_counts[prov] = counter
    for prov in providers_by_name:
        lim = providers_by_name[prov].get("limits", {}) or {}
        tr = lim.get("type_ranges", {}) or {}
        if not tr:
            continue
        counts = prov_type_counts.get(prov, collections.Counter())
        for t, rng in tr.items():
            if not isinstance(rng, (list, tuple)) or len(rng) != 2:
                continue
            mn, mx = rng[0], rng[1]
            cnt = counts.get(t, 0)
            if (mn is not None and cnt < mn) or (mx is not None and cnt > mx):
                type_range_viol.append((prov, t, cnt, mn, mx))
    add_check("Per-type min/max ranges respected", len(type_range_viol)==0, f"Violations: {len(type_range_viol)}")

    # ---- Print summary ----
    print(_c_head("\n=== Constraint Check Summary ==="), file=stream)
    for name, ok, details in checks:
        tag = _c_ok("[OK]") if ok else _c_fail("[FAIL]")
        print(f"{tag} {name}" + (f" — {details}" if details else ""), file=stream)

    def preview(label, rows, limit=8):
        if not rows: return
        print(_c_warn(f"\n{label} (showing up to {limit})"), file=stream)
        for r in rows[:limit]:
            print("  -", r, file=stream)
        if len(rows) > limit:
            print(f"  ... (+{len(rows)-limit} more)", file=stream)

    preview("Unfilled shifts", unfilled)
    preview("Overfilled shifts (shift_id, count)", overfilled)
    preview("Unknown shift IDs", unknown_shift_ids)
    preview("Unknown providers", unknown_providers)
    preview("Type-not-allowed violations (prov, shift_id, prov_type, allowed_types)", bad_allowed)
    preview("Forbidden-day violations (prov, date, shift_ids)", bad_forbidden)
    preview("Multiple shifts same day (prov, date, shift_ids)", multi_same_day)
    preview("Max-consecutive violations (prov, longest, K)", bad_consec)
    preview("Preferred-days HARD violations (prov, date, shift, shift_type, allowed_types)", bad_pref_hard)
    preview("Required-days HARD violations (prov, date, required_types, assigned_types)", hard_on_misses)
    preview("Min/max total violations (prov, total, min, max)", minmax_viol)
    preview("Per-type range violations (prov, type, count, min, max)", type_range_viol)

    # Soft-preference diagnostics (informational)
    print(_c_head("\n=== Soft-Preference Diagnostics (informational) ==="), file=stream)
    soft_off_hits = []
    soft_on_mismatch = []
    for prov, by_day in prov_day_to_shifts.items():
        p = providers_by_name.get(prov)
        if not p: continue
        soft_off = set(p.get("forbidden_days_soft", []))
        soft_on = p.get("preferred_days_soft", {}) or {}
        for d, sids in by_day.items():
            if d in soft_off:
                soft_off_hits.append((prov, d, sids))
            prefs = set(soft_on.get(d, []))
            if prefs and not any(shift_type_map.get(sid, "") in prefs for sid in sids):
                soft_on_mismatch.append((prov, d, [shift_type_map.get(sid, "") for sid in sids], sorted(prefs)))
    print(f"Worked on soft-off days: {len(soft_off_hits)}", file=stream)
    if soft_off_hits:
        for r in soft_off_hits[:preview_limit]:
            print("  -", r, file=stream)
        if len(soft_off_hits) > preview_limit:
            print(f"  ... (+{len(soft_off_hits)-preview_limit} more)", file=stream)
    print(f"Soft-on type mismatches: {len(soft_on_mismatch)}", file=stream)
    if soft_on_mismatch:
        for r in soft_on_mismatch[:preview_limit]:
            print("  -", r, file=stream)
        if len(soft_on_mismatch) > preview_limit:
            print(f"  ... (+{len(soft_on_mismatch)-preview_limit} more)", file=stream)

    # Cluster analysis
    print(_c_head("\n=== Cluster Analysis ==="), file=stream)
    prov_clusters = {}
    for prov, by_day in prov_day_to_shifts.items():
        prov_clusters[prov] = _cluster_sizes(by_day.keys())
    ranked = sorted(prov_clusters.items(), key=lambda kv: (len(kv[1]), sum(kv[1])), reverse=True)
    print("Providers ranked by number of clusters (then total worked days):", file=stream)
    for prov, sizes in ranked:
        print(f"  - {prov:20s}  clusters={len(sizes):2d}  sizes={sizes}", file=stream)

    # Imbalances
    print(_c_head("\n=== Imbalances ==="), file=stream)
    shifts_per_provider = {prov: sum(len(v) for v in by_day.values()) for prov, by_day in prov_day_to_shifts.items()}
    for name in providers_by_name:
        shifts_per_provider.setdefault(name, 0)
    total_assign = sum(shifts_per_provider.values())
    nprov = len(providers_by_name) or 1
    avg = total_assign / nprov
    dev = {prov: (cnt - avg) for prov, cnt in shifts_per_provider.items()}
    top_over = sorted(dev.items(), key=lambda kv: kv[1], reverse=True)[:10]
    top_under = sorted(dev.items(), key=lambda kv: kv[1])[:10]

    print(f"Total assignments: {total_assign} over {nprov} providers; average per provider: {avg:.2f}", file=stream)
    print("Top over-assigned (provider, +diff, total):", file=stream)
    for prov, d in top_over:
        print(f"  - {prov:20s}  +{d:.1f} (={shifts_per_provider[prov]})", file=stream)
    print("Top under-assigned (provider, -diff, total):", file=stream)
    for prov, d in top_under:
        print(f"  - {prov:20s}  {d:.1f} (=={shifts_per_provider[prov]})", file=stream)

    print("\nPer-provider counts by shift TYPE (nonzero only):", file=stream)
    all_types = sorted({t for t in (sh.get("type","") for sh in case["shifts"]) if t})
    prov_type_counts = {prov: collections.Counter() for prov in {p["name"] for p in case["providers"]}}
    for prov, by_day in prov_day_to_shifts.items():
        counter = collections.Counter()
        for d, sids in by_day.items():
            for sid in sids:
                counter[shift_type_map.get(sid, "")] += 1
        prov_type_counts[prov] = counter
    for prov in sorted(prov_type_counts.keys()):
        c = prov_type_counts.get(prov, collections.Counter())
        parts = [f"{t}:{c[t]}" for t in all_types if c[t] > 0]
        if parts:
            print(f"  - {prov}: " + ", ".join(parts), file=stream)

    overall_type_counts = collections.Counter(shift_type_map[sid] for sid in shifts_by_id if schedule_map.get(sid))
    print("\nOverall assigned counts by type:", file=stream)
    for t, cnt in overall_type_counts.most_common():
        print(f"  - {t or '(blank)'}: {cnt}", file=stream)

    # Weekday distribution
    print(_c_head("\n=== Weekday Distribution ==="), file=stream)
    weekday_counts = collections.Counter()
    for sid, provs in schedule_map.items():
        if sid not in shift_date: continue
        wd = iso_weekday_name(shift_date[sid])
        weekday_counts[wd] += len(provs)
    for wd in ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]:
        print(f"  - {wd:9s}: {weekday_counts.get(wd,0)}", file=stream)

    print(_c_dim("\nDone."), file=stream)

# -------------- File name sanitization --------------

def _sanitize_filename_part(s: str) -> str:
    s = "".join(ch if ch.isalnum() or ch in ("_", "-", ".") else "_" for ch in s.strip())
    if not s:
        s = "sheet"
    return s

# -------------- Public API (replaces CLI) --------------

def run_diag(case: str, schedule: str, *, no_color: bool = False, preview: int = 8) -> None:
    """
    Programmatic entry point mirroring the original CLI arguments.

    Args:
        case:      Path to testcase JSON (same as --case)
        schedule:  Path to schedule file (json/csv/xlsx/xlsm) (same as --schedule)
        no_color:  Disable ANSI colors in terminal (same as --no-color)
        preview:   Max items to preview per category (same as --preview)

    Side effects / outputs are identical to the CLI version:
      - Multi-sheet Excel: writes one .diagnose.txt per suitable sheet next to the schedule.
      - Single JSON/CSV (or single suitable sheet): prints the report to stdout.
    """
    global _USE_COLOR
    _USE_COLOR = (not no_color)

    case_obj = load_case(case)
    items, src = load_schedules(schedule)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(_c_head(f"(Loaded case encoding={case_obj['case_encoding']}; schedule source={src}; {ts})"))

    # Determine output directory and base name
    sched_path = Path(schedule)
    out_dir = sched_path.parent
    base = sched_path.stem

    # If multiple sheets, write one file per sheet. Also echo a short notice to terminal.
    if len(items) > 1 or str(src).startswith("xlsx"):
        written = []
        for label, sched_map in items:
            safe = _sanitize_filename_part(label)
            out_file = out_dir / f"{base}__{safe}.diagnose.txt"
            # force plain text (no ANSI) inside files
            use_color_prev = _USE_COLOR
            _USE_COLOR = False
            with open(out_file, "w", encoding="utf-8") as f:
                banner = f"=== DIAGNOSE: Sheet '{label}' from {sched_path.name} ==="
                f.write(f"{banner}\nGenerated: {ts}\n\n")
                diagnose(case_obj, sched_map, stream=f, preview_limit=preview)
            _USE_COLOR = use_color_prev
            written.append(out_file.name)
            print(_c_ok(f"[WROTE] {out_file}"))
        print(_c_head(f"\nDone. Wrote {len(written)} report file(s):"))
        for w in written:
            print("  -", w)
    else:
        # Single item (CSV/JSON or single-sheet Excel) -> print to stdout
        label, sched_map = items[0]
        banner = f"=== DIAGNOSE: {sched_path.name} ({label}) ==="
        print(_c_head(banner))
        diagnose(case_obj, sched_map, stream=sys.stdout, preview_limit=preview)

# ----------------------------- Logging helpers -----------------------------

class _StreamToLogger:
    """File-like object that redirects writes into a logger at a given level."""
    def __init__(self, logger: Logger, level: int):
        self.logger = logger
        self.level = level
        self._buffer = ""

    def write(self, buf):
        if not isinstance(buf, str):
            buf = buf.decode("utf-8", errors="replace")
        self._buffer += buf
        while "\n" in self._buffer:
            line, self._buffer = self._buffer.split("\n", 1)
            line = line.rstrip()
            if line:
                self.logger.log(self.level, line)

    def flush(self):
        if self._buffer:
            self.logger.log(self.level, self._buffer.rstrip())
            self._buffer = ""

# ----------------------------- Logging helpers -----------------------------
# Drop-in GUI logging handler that is thread-safe via a Queue + periodic drain.

class TkQueueHandler(logging.Handler):
    """Enqueue log records; GUI drains them on the Tk main thread."""
    def __init__(self, gui):
        super().__init__()
        self.gui = gui  # TestcaseGUI instance

    def emit(self, record):
        try:
            msg = self.format(record)
            # Never touch Tk here (this may run on worker threads).
            self.gui._log_queue.put(msg)
        except Exception:
            pass

def _mk_logger(out_dir:str, ts:str) -> Logger:
    """
    Create a console + file logger at INFO level.
    Also safe to redirect sys.stdout/sys.stderr to it (we log to sys.__stdout__).
    """
    os.makedirs(out_dir, exist_ok=True)
    log_path = os.path.join(out_dir, f"scheduler_run.log")

    logger = logging.getLogger("scheduler")
    logger.setLevel(logging.INFO)
    logger.propagate = False

    # Keep any existing TkQueueHandler (GUI), remove everything else.
    for h in list(logger.handlers):
        if not isinstance(h, TkQueueHandler):
            logger.removeHandler(h)

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    ch = logging.StreamHandler(sys.__stdout__)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(fmt)

    logger.addHandler(ch)
    logger.addHandler(fh)
    logger.info("Logging to %s", log_path)
    return logger

# --------------------------------------------------------------------------

def day_name(iso_date: str) -> str:
    y,m,d = map(int, iso_date.split('-'))
    return dt.date(y,m,d).strftime('%A')

def parse_iso_minutes(ts: str) -> int:
    return int(dt.datetime.fromisoformat(ts).timestamp() // 60)

def safe_get(d, *keys, default=None):
    cur = d
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur

def get_num(d, *keys, default=0.0):
    v = safe_get(d, *keys, default=None)
    try:
        return float(v) if v is not None else float(default)
    except (TypeError, ValueError):
        return float(default)

def _extract_numeric(consts, path, default):
    """Extract a numeric from nested dict using dotted path.
    Returns (used_value, raw_value, default, defaulted_flag)."""
    keys = path.split('.')
    cur = consts
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            raw = None
            break
        cur = cur[k]
    else:
        raw = cur
    try:
        if raw is None:
            used = float(default); defaulted = True
        else:
            used = float(raw); defaulted = False
    except (TypeError, ValueError):
        used = float(default); defaulted = True
    return used, raw, default, defaulted

def _write_constants_log(consts, out_dir, cli_time_override=None):
    # what we log
    numeric_paths = [
        ("weights.hard.uncovered_shift", 0.0),
        ("weights.hard.slack_unfilled", 0.0),
        ("weights.hard.slack_shift_less", 0.0),
        ("weights.hard.slack_shift_more", 0.0),
        ("weights.hard.slack_cant_work", 0.0),
        ("weights.hard.slack_consec", 0.0),
        ("weights.hard.rest_12h", 0.0),
        ("weights.hard.type_range", 0.0),
        ("weights.hard.weekend_range", 0.0),
        ("weights.hard.total_limit", 0.0),
        ("weights.hard.consecutive", 0.0),
        ("weights.soft.weekday_pref", 1.0),
        ("weights.soft.type_pref", 1.0),
        ("weights.soft.cluster_size", 15.0),
        ("weights.soft.requested_off", 3.0),
        ("weights.soft.days_wanted_not_met", 0.0),
        ("weights.soft.cluster_any_start", 0.0),
        ("weights.soft.cluster_weekend_start", 0.0),
        ("weights.soft.transitions_any", 0.0),
        ("weights.soft.transitions_night", 0.0),
        ("objective.hard", 1.0),
        ("objective.soft", 1.0),
        ("objective.fair", 0.0),
        ("solver.max_time_in_seconds", 120.0),
        ("solver.phase1_fraction", 0.4),
        ("solver.relative_gap", 0.01),
        ("solver.num_threads", 8),
    ]
    flags = {
        "solver.min_total_is_hard": bool(((consts.get("solver") or {}).get("min_total_is_hard", False)))
    }
    rows = []
    defaulted = []
    for path, default in numeric_paths:
        used, raw, dflt, was_defaulted = _extract_numeric(consts, path, default)
        rows.append({"path": path, "raw": raw, "used": used, "default": dflt, "defaulted": was_defaulted})
        if was_defaulted:
            defaulted.append(path)

    # apply CLI overrides to the report (not mutating consts here)
    if cli_time_override is not None:
        for r in rows:
            if r["path"] == "solver.max_time_in_seconds":
                r["used"] = float(cli_time_override)
                r["note"] = "overridden_by_run_config"

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(out_dir, exist_ok=True)
    # JSON
    jpath = os.path.join(out_dir, f"constants_effective.json")
    with open(jpath, "w", encoding="utf-8") as f:
        json.dump({"values": rows, "flags": flags}, f, indent=2)
    # TXT
    tpath = os.path.join(out_dir, f"constants_effective.log")
    with open(tpath, "w", encoding="utf-8") as f:
        f.write("Effective constants (after null/default handling)\n")
        if cli_time_override is not None:
            f.write(f"(time overridden by case.run.time to {cli_time_override} sec)\n")
        for r in rows:
            mark = " *DEFAULTED* " if r["defaulted"] else ""
            note = f" [{r.get('note')}]" if r.get("note") else ""
            f.write(f"- {r['path']}: used={r['used']} raw={r['raw']} default={r['default']}{mark}{note}\n")
        f.write("\nFlags:\n")
        for k,v in flags.items():
            f.write(f"- {k}: {v}\n")
        if defaulted:
            f.write("\nWARNING: The following paths were null/invalid and defaulted:\n")
            for p in defaulted:
                f.write(f"  - {p}\n")
    print("Wrote:", tpath)
    return tpath, jpath

def norm_name(name: str) -> str:
    return re.sub(r'[^a-z]', '', (name or '').lower())

def infer_allowed_types(shift: Dict[str,Any], provider_types: List[str]) -> set:
    """Allowed types for a shift."""
    allowed = shift.get('allowed_provider_types')
    if allowed:
        return set(allowed)
    stype = shift.get('type', '')
    if isinstance(stype, str) and '_' in stype:
        prefix = stype.split('_', 1)[0]
        if prefix in provider_types:
            return {prefix}
    return set(provider_types)

def merge_case_limits(case: Dict[str,Any]) -> Dict[str,Any]:
    """Merge optional top-level case['limits'] into per-provider limits by fuzzy name."""
    if 'limits' not in case or not isinstance(case['limits'], dict):
        return case
    lims = case['limits']
    provs = case.get('providers', [])
    def findp(n):
        target = norm_name(n)
        for p in provs:
            if norm_name(p.get('name','')) == target:
                return p
        toks = [t for t in re.split(r'\s+', (n or '').lower()) if t]
        for p in provs:
            nrm = norm_name(p.get('name',''))
            if all(t in nrm for t in toks):
                return p
        return None

    for it in lims.get('total', []):
        p = findp(it.get('provider',''))
        if not p: continue
        p.setdefault('limits', {})
        rng = it.get('range', [])
        if len(rng)==2:
            p['limits']['min_total'] = int(rng[0])
            p['limits']['max_total'] = int(rng[1])

    for it in lims.get('weekend', []):
        p = findp(it.get('provider',''))
        if not p: continue
        p.setdefault('limits', {})
        rng = it.get('range', [])
        if len(rng)==2:
            p['limits']['weekend_range'] = [int(rng[0]), int(rng[1])]

    for it in lims.get('consecutive', []):
        p = findp(it.get('provider',''))
        if not p: continue
        p.setdefault('limits', {})
        cap = it.get('max', None)
        if cap is not None:
            p['limits']['max_consecutive_days'] = int(cap)

    for it in lims.get('type_ranges', []):
        p = findp(it.get('provider',''))
        if not p: continue
        p.setdefault('limits', {})
        tr = p['limits'].setdefault('type_ranges', {})
        t = it.get('type', None)
        rng = it.get('range', [])
        if t and len(rng)==2:
            tr[t] = [int(rng[0]), int(rng[1])]

    return case



def load_inputs_from_case(case_path: str):
    """Merged format: a single JSON file that contains both the case and a 'constants' section,
    plus optional 'run' section for output path, k, seed, and total time."""
    with open(case_path, 'r', encoding='utf-8') as f:
        case = json.load(f)

    # Inline constants inside the case
    consts = case.get('constants', {}) or {}

    # Optional legacy pointer to external constants
    cpath = case.get('constants_path')
    if cpath and isinstance(cpath, str) and len(cpath) > 0:
        try:
            with open(cpath, 'r', encoding='utf-8') as g:
                consts = json.load(g)
        except Exception:
            pass  # ignore if not found

    # Merge case-level limits into providers by name (same as before)
    case = merge_case_limits(case)

    # Normalize weekend defaults if missing
    case.setdefault('calendar', {})
    case['calendar'].setdefault('weekend_days', (consts.get('calendar', {}) or {}).get('weekend_days', ['Saturday','Sunday']))

    # Normalize provider schema fields (safe defaults)
    for p in case.get("providers", []):
        p.setdefault("forbidden_days_hard", [])
        p.setdefault("preferred_days_hard", {})
        p.setdefault("preferred_days_soft", {})
        if not isinstance(p.get("max_consecutive_days", None), int):
            p["max_consecutive_days"] = 1000

    return consts, case

def load_inputs(constants_path: str, case_path: str):
    with open(constants_path, 'r', encoding='utf-8') as f:
        consts = json.load(f)
    with open(case_path, 'r', encoding='utf-8') as f:
        case = json.load(f)
    case = merge_case_limits(case)
    case.setdefault('calendar', {})
    case['calendar'].setdefault('weekend_days', consts.get('calendar', {}).get('weekend_days', ['Saturday','Sunday']))
    return consts, case

from ortools.sat.python import cp_model

def build_model(consts: Dict[str,Any], case: Dict[str,Any]) -> Dict[str,Any]:
    logger = logging.getLogger("scheduler")

    days: List[str] = case['calendar']['days']
    weekend_names: List[str] = case['calendar']['weekend_days']
    weekend_idx = {i for i,d in enumerate(days) if day_name(d) in weekend_names}

    shifts: List[Dict[str,Any]] = case['shifts']
    providers: List[Dict[str,Any]] = case['providers']
    S = list(range(len(shifts)))
    P = list(range(len(providers)))
    D = list(range(len(days)))

    provider_types = sorted(set([p.get('type','MD') for p in providers]))

    date_to_idx = {d:i for i,d in enumerate(days)}
    shift_day = [date_to_idx[shifts[s]['date']] for s in S]
    shift_type = [shifts[s]['type'] for s in S]
    types = sorted(set(shift_type))
    type_to_idx = {t:i for i,t in enumerate(types)}
    day_to_shifts = {d: [] for d in D}
    for s in S: day_to_shifts[shift_day[s]].append(s)

    # --- Log high-level instance stats (read-only) ---
    logger.info("Instance: |days|=%d |shifts|=%d |providers|=%d |types|=%d", len(D), len(S), len(P), len(types))
    logger.info("Weekend names: %s", weekend_names)
    logger.info("Weekend indices: %s", sorted(list(weekend_idx)))
    if shifts:
        logger.info("Shift[0] keys: %s", list(shifts[0].keys()))
    if providers:
        logger.info("Provider[0] keys: %s", list(providers[0].keys()))
    logger.info("Provider types: %s", provider_types)

    model = cp_model.CpModel()

    x = {(i, j): model.NewBoolVar(f"x_{i}_{j}") for i in S for j in P}
    slack_unfilled = [model.NewBoolVar(f"slack_{i}_unfilled") for i in S]
    for i in S:
        model.Add(sum(x[i, j] for j in P) + slack_unfilled[i] == 1)

    # Max consective days
    max_consec = [providers[j].get('max_consecutive_days', 0) for j in P]
    max_consec = [0 if not i else i for i in max_consec]
    logger.info("max_consecutive_days per provider (0 means no cap): %s", max_consec)

    # --- Slack for max consecutive days: slack = max(0, max_cluster - max_consec) ---

# ===== Longest consecutive working-day cluster and slack (exact violation size) =====

    N = len(D)

    # y[(i,d)] = 1 iff provider i works any shift on day d
    y = {}
    for i in P:
        for d in D:
            yi = model.NewBoolVar(f"workday_{i}_{d}")
            y[(i, d)] = yi
            Sd = day_to_shifts.get(d, [])
            if not Sd:
                model.Add(yi == 0)
            else:
                # If yi = 0 -> all x[s,i] = 0 ; If yi = 1 -> at least one x[s,i] = 1
                for s in Sd:
                    model.Add(x[(s, i)] <= yi)
                model.Add(sum(x[(s, i)] for s in Sd) >= yi)

    # run[d] = current consecutive streak length ending at day d
    max_clusters = [model.NewIntVar(0, N, f"max_cluster_{i}") for i in P]
    runs = {}  # store for reuse in cubes
    for i in P:
        run = [model.NewIntVar(0, N, f"run_{i}_{d}") for d in D]
        runs[i] = run
        # Day 0
        model.Add(run[0] == 1).OnlyEnforceIf(y[(i, 0)])
        model.Add(run[0] == 0).OnlyEnforceIf(y[(i, 0)].Not())
        # Days 1..N-1
        for d in range(1, N):
            yt, yp = y[(i, d)], y[(i, d - 1)]
            model.Add(run[d] == 0).OnlyEnforceIf(yt.Not())                 # not working today -> reset
            model.Add(run[d] == 1).OnlyEnforceIf([yt, yp.Not()])           # start new streak
            model.Add(run[d] == run[d - 1] + 1).OnlyEnforceIf([yt, yp])    # extend streak
        model.AddMaxEquality(max_clusters[i], run)

    # ----- Slack: slack_consec[i] = max(0, max_clusters[i] - max_consec[i]) -----
    _zero = model.NewIntVar(0, 0, "zero_const")
    slack_consec = [model.NewIntVar(0, N, f"cons_slack_{i}") for i in P]
    for i in P:
        if max_consec[i] > 0:
            diff = model.NewIntVar(-N, N, f"cluster_overrun_{i}")
            model.Add(diff == max_clusters[i] - max_consec[i])
            model.AddMaxEquality(slack_consec[i], [diff, _zero])
            model.Add(max_consec[i] - max_clusters[i] + slack_consec[i] >= 0)
        else:
            model.Add(slack_consec[i] == 0)

    # ----- NEW: Cubic penalty of cluster lengths per provider (soft) -------------
    # Detect cluster ends: end_d = 1 iff y[i,d]==1 and (d==N-1 or y[i,d+1]==0)
    # Cluster length is runs[i][d_end]; we gate it into L_d and build L^3.
    cluster_cubesums = [model.NewIntVar(0, N**3, f"cluster_cubesum_{i}") for i in P]

    for i in P:
        cube_terms = []
        for d in D:
            end_d = model.NewBoolVar(f"cluster_end_{i}_{d}")
            if d < N - 1:
                # Linearization of end_d == (y_d == 1 and y_{d+1} == 0)
                model.Add(end_d >= y[(i, d)] - y[(i, d + 1)])
                model.Add(end_d <= y[(i, d)])
                model.Add(end_d <= 1 - y[(i, d + 1)])
            else:
                # Last day: end iff working that day
                model.Add(end_d == y[(i, d)])

            # L_d = run[d] if end_d else 0
            Ld = model.NewIntVar(0, N, f"cluster_len_{i}_{d}")
            model.Add(Ld == runs[i][d]).OnlyEnforceIf(end_d)
            model.Add(Ld == 0).OnlyEnforceIf(end_d.Not())

            # L^2 and L^3
            L2 = model.NewIntVar(0, N * N, f"cluster_len_sq_{i}_{d}")
            model.AddMultiplicationEquality(L2, [Ld, Ld])
            L3 = model.NewIntVar(0, N * N * N, f"cluster_len_cube_{i}_{d}")
            model.AddMultiplicationEquality(L3, [L2, Ld])

            cube_terms.append(L3)

        if cube_terms:
            model.Add(cluster_cubesums[i] == sum(cube_terms))
        else:
            model.Add(cluster_cubesums[i] == 0)
    # ---------------------------------------------------------------------------

    # 12 hrs apart
    for iter1 in S:
        for iter2 in S:
            if shifts[iter1]["id"] == shifts[iter2]["id"]:
                continue

            # Remove timezone info for safe comparison
            s1_str = shifts[iter1]["start"].replace('Z', '').split('+')[0]
            s2_str = shifts[iter2]["start"].replace('Z', '').split('+')[0]
            e1_str = shifts[iter1]["end"].replace('Z', '').split('+')[0]
            
            s1 = dt.datetime.fromisoformat(s1_str)
            s2 = dt.datetime.fromisoformat(s2_str)
            e1 = dt.datetime.fromisoformat(e1_str)

            if s1 > s2:
                continue
            
            # Check if shifts overlap or are less than 12 hours apart
            is_too_close = False
            if s2 < e1:  # Direct overlap
                is_too_close = True
            else:
                diff = s2 - e1
                if diff.total_seconds() < 12 * 3600:
                    is_too_close = True

            if is_too_close:
                for j in P:
                    model.AddAtMostOne([x[iter1, j], x[iter2, j]])
    # cant because type
    for s in S:
        for p in P:
            if providers[p].get('type') not in shifts[s]["allowed_provider_types"]:
                model.Add(x[s, p] == 0)

    # provider hard limits, we are trying to minimize slacks
    slack_shift_less = [model.NewIntVar(0, 1000, f"shifts_{j}") for j in P]
    slack_shift_more = [model.NewIntVar(0, 1000, f"shifts_{j}") for j in P]

    slack_hard_yes = [model.NewIntVar(0, 1000, f"shifts_{j}") for j in P]

    for j in P:
        print(j)
        lim = providers[j].get('limits', {}) or {}
        min_total = int(lim.get('min_total', 0))
        max_total = int(lim.get('max_total', len(S)))
        model.Add(sum(x[i, j] for i in S) + slack_shift_less[j] >= min_total)
        model.Add(sum(x[i, j] for i in S) - slack_shift_more[j] <= max_total)

    #respect days that a provider cant
    slack_cant_work = [model.NewIntVar(0, len(S), f"cantwork_{j}") for j in P]
    for j in P:
        forb = set(providers[j].get('forbidden_days_hard', []))
        terms = [x[s, j] for s in S if shifts[s]['date'] in forb]
        if terms:
            model.Add(slack_cant_work[j] == sum(terms))
        else:
            model.Add(slack_cant_work[j] == 0)


    # -------- HARD "ON" (date -> specific shift types) slack --------
    # Uses same coefficient as hard OFF in the hard objective.
    slack_hard_on = [model.NewIntVar(0, 1000, f"hardon_{j}") for j in P]
    ANY = "ANY"
    for j in P:
        terms = []
        hard_on_map = (providers[j].get('preferred_days_hard') or {})
        for d_str, tlist in hard_on_map.items():
            if d_str not in date_to_idx:
                continue
            if not tlist:
                # no specific types requested => no requirement
                continue
            d = date_to_idx[d_str]
            Sh = day_to_shifts.get(d, [])
            if not Sh:
                miss = model.NewBoolVar(f"hard_on_miss_{j}_{d}")
                model.Add(miss == 1)
                terms.append(miss)
                continue
            R = Sh if (ANY in tlist) else [s for s in Sh if shift_type[s] in set(tlist)]
            if not R:
                miss = model.NewBoolVar(f"hard_on_miss_{j}_{d}")
                model.Add(miss == 1)
                terms.append(miss)
                continue
            sel = model.NewBoolVar(f"hard_on_sel_{j}_{d}")
            for s in R:
                model.Add(x[s, j] <= sel)
            model.Add(sum(x[s, j] for s in R) >= sel)
            miss = model.NewBoolVar(f"hard_on_miss_{j}_{d}")
            model.Add(sel + miss == 1)
            terms.append(miss)
        if terms:
            model.Add(slack_hard_on[j] == sum(terms))
        else:
            model.Add(slack_hard_on[j] == 0)

    # Objective: minimize total slack
    c_slack_unfilled = int(get_num(consts, 'weights', 'hard', 'slack_unfilled', default=1))
    c_slack_shift_less = int(get_num(consts, 'weights', 'hard', 'slack_shift_less', default=1))
    c_slack_shift_more = int(get_num(consts, 'weights', 'hard', 'slack_shift_more', default=1))
    c_slack_cant_work = int(get_num(consts, 'weights', 'hard', 'slack_cant_work', default=1))
    c_slack_consec = int(get_num(consts, 'weights', 'hard', 'slack_consec', default=1))

    U = model.NewIntVar(0, 10**18, "U")
    model.Add(U == c_slack_unfilled * sum(slack_unfilled) 
              + c_slack_shift_less * sum(slack_shift_less) 
              + c_slack_shift_more * sum(slack_shift_more) 
              + c_slack_cant_work * sum(slack_cant_work) 
              + c_slack_consec * sum(slack_consec)
              + c_slack_cant_work * sum(slack_hard_on))  # NEW: hard ON slack weighted like hard OFF
    model.minimize(U)

    # Phase-1 solve (hard slacks) — VERBOSE + callback into logger
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(120)
    solver.parameters.num_search_workers = 8
    solver.parameters.log_search_progress = True
    solver.parameters.log_to_stdout = False
    try:
        solver.log_callback = lambda line: logging.getLogger("scheduler").info("[phase1] %s", line.rstrip())
    except Exception:
        pass
    logger.info("Phase-1 solve: time=%ss workers=%s", solver.parameters.max_time_in_seconds, solver.parameters.num_search_workers)

    st1 = solver.Solve(model)
    obj1 = solver.ObjectiveValue() if st1 in (cp_model.OPTIMAL, cp_model.FEASIBLE) else None
    logger.info("Phase-1 status=%s objective(U)=%s", solver.StatusName(st1), obj1)
    # Optional: this goes through the logger anyway
    print(solver.ObjectiveValue() if obj1 is not None else "No objective")


    # slacks enforced
    # now we solve for soft constraints
    for i in (slack_unfilled + slack_shift_less + slack_shift_more + slack_cant_work + slack_consec + slack_hard_on):
        model.Add(i == solver.Value(i))

    # soft penalty 
    Weighted = model.NewIntVar(0, 1000000000000000000, "Weighted")

    shifts_by_type = {t: [s for s in S if shift_type[s] == t] for t in types}
    for t in shifts_by_type:
        shifts_by_type[t].sort(key=lambda s: shift_day[s])

    # Cluster starts per provider j, per type t, along the sorted shift sequence
    cluster_start = {}
    cluster_count = {}

    #starting a cluster in 
    for j in P:
        for t in types:
            seq = shifts_by_type[t]
            if not seq:
                continue

            # first of each type starts a cluster if taken
            s0 = seq[0]
            cluster_start[(j, t, 0)] = model.NewBoolVar(f"cluster_start_{j}_{type_to_idx[t]}_0")
            model.Add(cluster_start[(j, t, 0)] == x[s0, j])

            # new cluster when previous not taken and current taken: 0 -> 1 transition
            for k in range(1, len(seq)):
                sp, sc = seq[k - 1], seq[k]
                vk = model.NewBoolVar(f"cluster_start_{j}_{type_to_idx[t]}_{k}")
                cluster_start[(j, t, k)] = vk
                model.Add(vk >= x[sc, j] - x[sp, j])
                model.Add(vk <= x[sc, j])
                model.Add(vk <= 1 - x[sp, j])

            # total clusters for (j, t)
            cc_jt = model.NewIntVar(0, len(seq), f"cluster_count_{j}_{type_to_idx[t]}")
            model.Add(cc_jt == sum(cluster_start[(j, t, k)] for k in range(len(seq))))
            cluster_count[(j, t)] = cc_jt

    # === Provider-level cluster counts ===
    max_clusters_per_provider = sum(len(shifts_by_type[t]) for t in types)
    cc = [model.NewIntVar(0, max_clusters_per_provider, f"cc_{j}") for j in P]
    for j in P:
        model.Add(cc[j] == sum(cluster_count[(j, t)] for t in types if (j, t) in cluster_count))

    # Personal dissatisfaction is:
    # - number of clusters (any type)
    # - Violations of soft requirements

    days_per_provider = [model.NewIntVar(0, 40, f"days_per_provider_{i}") for i in P]
    for i in P:
        model.Add(days_per_provider[i] == sum([x[j, i] for j in S]))
    clusters_per_provider = [model.NewIntVar(0, 10**15, f"personal_penalty_{j}") for j in P]
    for p in P:
        model.Add(clusters_per_provider[p] == cc[p])
    cluster_square = [model.NewIntVar(0, 10**5, f"cluster_square_{j}") for j in P]
    for p in P:
        model.AddMultiplicationEquality(cluster_square[p], [clusters_per_provider[p], clusters_per_provider[p]])
    nshifts = len(S)
    nproviders = len(P)
    avg = nshifts // nproviders

    nshifts = len(S)

    # slacks (use distinct names!)
    slack_less = [model.NewIntVar(0, nshifts, f"less_shifts_s_{i}") for i in P]
    slack_more = [model.NewIntVar(0, nshifts, f"more_shifts_s_{i}") for i in P]

    # tie slacks to deviation from avg
    for i in P:
        model.Add(sum(x[s, i] for s in S) + slack_less[i] >= avg)
        model.Add(sum(x[s, i] for s in S) - slack_more[i] <= avg)

    # square the slacks via auxiliary vars
    less_sq = [model.NewIntVar(0, 2 *nshifts * nshifts, f"less_sq_{i}") for i in P]
    more_sq = [model.NewIntVar(0, 2 * nshifts * nshifts, f"more_sq_{i}") for i in P]
    for i in P:
        model.AddMultiplicationEquality(less_sq[i], [slack_less[i], slack_less[i]])
        model.AddMultiplicationEquality(more_sq[i], [slack_more[i], slack_more[i]])

    # deviation[i] = less_sq[i] + more_sq[i]
    deviation = [model.NewIntVar(0, 2 * nshifts * nshifts, f"deviation_{i}") for i in P]
    cclusters = int(get_num(consts, 'weights', 'soft', 'cluster', default=500))
    cunfair = int(get_num(consts, 'weights', 'soft', 'unfair_number', default=10))
    cweekend_not_clustered = int(get_num(consts, 'weights', 'soft', 'cluster_weekend_start', default=50000))
    c_soft_on = int(get_num(consts, 'weights', 'soft', 'days_wanted_not_met', default=10))
    c_soft_off = int(get_num(consts, 'weights', 'soft', 'requested_off', default=10))
    c_cluster_size = int(get_num(consts, 'weights', 'soft', 'cluster_size', default=10))
    #It would be better to add a new cluster than one of size n if n^3 * c_cluster_size > cclusters
    import math
   # c_cluster_size = int(math.ceil((cclusters / c_cluster_size) ** 0.666))
    for i in P:
        model.Add(deviation[i] == less_sq[i] + more_sq[i])    
    # Weekend penalty (works Sat, not Sun)
    second_weekend = max(weekend_idx)
    count_horrible = [model.NewIntVar(0, nshifts, f'weekend_unclustered_{i}') for i in P]

    import datetime as _dt

    D = range(len(days))
    def _wd(d):  # weekday: Mon=0..Sun=6
        y, m, dd = map(int, days[d].split('-'))
        return _dt.date(y, m, dd).weekday()

    weekend_pairs = [(d, d+1) for d in range(len(days)-1) if _wd(d) == 5 and _wd(d+1) == 6]

    y = {}  # (i,d) -> BoolVar
    for i in P:
        for d in D:
            yi = model.NewBoolVar(f"works_day_{i}_{d}")
            y[(i, d)] = yi
            S_d = day_to_shifts[d]
            if not S_d:
                model.Add(yi == 0)
            else:
                for s in S_d:
                    model.Add(x[s, i] <= yi)
                model.Add(sum(x[s, i] for s in S_d) >= yi)

    count_horrible = [model.NewIntVar(0, len(weekend_pairs), f"wk_pen_count_{i}") for i in P]
    for i in P:
        terms = []
        for d1, d2 in weekend_pairs:
            diff = model.NewIntVar(-1, 1, f"wk_diff_%d_%d_%d" % (i, d1, d2))
            pen  = model.NewIntVar(0, 1,  f"wk_pen_%d_%d_%d"  % (i, d1, d2))
            model.Add(diff == y[(i, d1)] - y[(i, d2)])
            model.AddMaxEquality(pen, [diff, model.NewConstant(0)])
            terms.append(pen)
        model.Add(count_horrible[i] == sum(terms))

    ANY = "ANY"  # keep if you want the same sentinel everywhere

    soft_off_i = [model.NewIntVar(0, len(days), f"soft_off_{i}") for i in P]
    soft_on_i  = [model.NewIntVar(0, len(days), f"soft_on_{i}")  for i in P]

    for i in P:
        # --- SOFT OFF (uses list of dates) ---
        off_terms = []
        soft_off_days = set(providers[i].get('forbidden_days_soft', []))
        for d_str in soft_off_days:
            if d_str not in date_to_idx:
                continue
            d = date_to_idx[d_str]
            Sh = day_to_shifts.get(d, [])
            if not Sh:
                continue  # nothing to avoid that day
            viol = model.NewBoolVar(f"soft_off_viol_{i}_{d}")
            # viol == 1 iff provider i works any shift that day
            for s in Sh:
                model.Add(x[s, i] <= viol)
            model.Add(sum(x[s, i] for s in Sh) >= viol)
            off_terms.append(viol)
        model.Add(soft_off_i[i] == (sum(off_terms) if off_terms else 0))

        # --- SOFT ON (unchanged, already correct) ---
        on_miss_terms = []
        soft_on_map  = providers[i].get('preferred_days_soft', {})
        for d_str, tlist in soft_on_map.items():
            if d_str not in date_to_idx:
                continue
            d = date_to_idx[d_str]
            Sh = day_to_shifts.get(d, [])
            if not Sh:
                continue
            R = Sh if (ANY in tlist) else [s for s in Sh if shift_type[s] in set(tlist)]
            if not R:
                continue
            sel  = model.NewBoolVar(f"soft_on_sel_{i}_{d}")
            for s in R:
                model.Add(x[s, i] <= sel)
            model.Add(sum(x[s, i] for s in R) >= sel)
            miss = model.NewBoolVar(f"soft_on_miss_{i}_{d}")
            model.Add(sel + miss == 1)
            on_miss_terms.append(miss)
        model.Add(soft_on_i[i] == (sum(on_miss_terms) if on_miss_terms else 0))


    model.Add(Weighted == cclusters * sum(cluster_square) +
                            cunfair * sum(deviation) +
                            c_cluster_size * sum(cluster_cubesums) +   # <<< NEW TERM
                            cweekend_not_clustered * sum(count_horrible) + 
                            c_soft_on * sum(soft_on_i) + 
                            c_soft_off * sum(soft_off_i))
    print(count_horrible)
    model.minimize(Weighted)
    # (Phase-2 solver is created in solve_two_phase)
    return dict(
        model=model,
        x=x,
        U=U,
        Weighted=Weighted,
        days=days,
        providers=providers,
        shifts=shifts,
        S=S, P=P, D=D,
        weekend_idx=weekend_idx,
        shift_day=shift_day,
        shift_type=shift_type,
        types=types,
        type_to_idx=type_to_idx,
        day_to_shifts=day_to_shifts,
    )

class KeepTopK(cp_model.CpSolverSolutionCallback):
    def __init__(self, x, K, days, providers, shifts):
        super().__init__()
        self.x = x; self.K=max(1,int(K))
        self.days=days; self.providers=providers; self.shifts=shifts
        self.pool=[]; self.seen=set()

    def on_solution_callback(self):
        # Iterate only over existing decision vars (s,j) in sparse x
        assign = [ (s,j) for (s,j), var in self.x.items() if self.Value(var) == 1 ]
        key = tuple(sorted(assign))
        if key in self.seen: 
            return
        self.seen.add(key)
        meta = {
            "objective": self.ObjectiveValue(),
            "best_bound": self.BestObjectiveBound(),
            "conflicts": self.NumConflicts(),
            "branches": self.NumBranches(),
            "wall_time_s": self.WallTime(),
            "assignments": len(assign),
        }
        self.pool.append((self.ObjectiveValue(),
                          {"assignment": key, "days": self.days, "providers": self.providers, "shifts": self.shifts},
                          meta))
        self.pool.sort(key=lambda z: z[0])
        if len(self.pool) > self.K:
            self.pool.pop()

class AssignmentPoolCollector(cp_model.CpSolverSolutionCallback):
    """Collect many assignment solutions (and their flattened 0/1 vectors) in a single run."""
    def __init__(self, x, S, P, days, providers, shifts, *, sense='min', obj_slack=None, pool_limit=20000, dedup=True):
        super().__init__()
        self.x = x                          # dict[(s,j)] -> BoolVar (sparse!)
        self.S = list(S)
        self.P = list(P)
        self.days = days
        self.providers = providers
        self.shifts = shifts
        self.sense = sense
        self.obj_slack = obj_slack
        self.pool_limit = pool_limit
        self.dedup = dedup
        self.pool = []        # [(obj, table, meta)]
        self.pool_vecs = []   # [tuple of 0/1]
        self._seen_vecs = set()
        self._best = None

    def _pack_vec(self):
        """Dense (S x P) bitvector; missing x[(s,j)] are treated as 0 without calling Value()."""
        vec = []
        get = self.x.get
        for s in self.S:
            for j in self.P:
                var = get((s, j))
                if var is None:
                    vec.append(0)
                else:
                    vec.append(1 if self.Value(var) == 1 else 0)
        return tuple(vec)

    def on_solution_callback(self):
        obj = self.ObjectiveValue()
        if self._best is None:
            self._best = obj
        else:
            if self.sense == 'min':
                if obj < self._best: self._best = obj
            else:
                if obj > self._best: self._best = obj

        if self.obj_slack is not None:
            if self.sense == 'min' and obj > self._best + self.obj_slack:
                return
            if self.sense == 'max' and obj < self._best - self.obj_slack:
                return

        vec = self._pack_vec()
        if self.dedup and vec in self._seen_vecs:
            return

        # Build assignment list only over existing (s,j) keys
        assign = [(s,j) for (s,j), var in self.x.items() if self.Value(var) == 1]
        key = tuple(sorted(assign))
        meta = {
            "objective": obj,
            "best_bound": self.BestObjectiveBound(),
            "conflicts": self.NumConflicts(),
            "branches": self.NumBranches(),
            "wall_time_s": self.WallTime(),
            "assignments": len(assign),
        }
        table = {"assignment": key, "days": self.days, "providers": self.providers, "shifts": self.shifts}

        # Pool management
        if self.pool_limit is not None and len(self.pool) >= self.pool_limit:
            if self.sense == 'min':
                worst_idx = max(range(len(self.pool)), key=lambda k: self.pool[k][0])
                if obj >= self.pool[worst_idx][0]:
                    return
            else:
                worst_idx = min(range(len(self.pool)), key=lambda k: self.pool[k][0])
                if obj <= self.pool[worst_idx][0]:
                    return
            del self.pool[worst_idx]
            del self.pool_vecs[worst_idx]

        self.pool.append((obj, table, meta))
        self.pool_vecs.append(vec)
        if self.dedup:
            self._seen_vecs.add(vec)

def _hamming(a: tuple, b: tuple) -> int:
    return sum(1 for u, v in zip(a, b) if u != v)

def _select_diverse_k(cb_pool, cb_vecs, K: int, L: int, *, sense='min', relax_to: int = 0):
    """Greedy best-first by objective; keep solutions whose Hamming distance to all kept ≥ L."""
    idxs = list(range(len(cb_pool)))
    if sense == 'min':
        idxs.sort(key=lambda k: cb_pool[k][0])  # obj asc
    else:
        idxs.sort(key=lambda k: -cb_pool[k][0]) # obj desc

    selected = []
    thr = int(L or 0)
    while True:
        for k in idxs:
            vec = cb_vecs[k]
            if all(_hamming(vec, cb_vecs[j]) >= thr for j in selected):
                selected.append(k)
                if len(selected) == K:
                    return selected
        if thr > relax_to:
            thr -= 1
        else:
            return selected
# ------------------------------------------------------------------------------------

def solve_two_phase(consts, case, ctx, K, seed=None):
    logger = logging.getLogger("scheduler")

    total_time = float(get_num(consts, 'solver', 'max_time_in_seconds', default=120))
    frac = float(get_num(consts, 'solver', 'phase1_fraction', default=0.4))
    t1 = max(5.0, total_time * frac)
    t2 = max(5.0, total_time - t1)
    logger.info("Time budget total=%.2fs split: phase1=%.2fs phase2=%.2fs", total_time, t1, t2)

    # Phase-2 directly on ctx['model'] with soft objective (existing pipeline)
    ctx2 = ctx
    model2 = ctx2['model']

    sp = consts.get('solver', {})
    solver2 = cp_model.CpSolver()
    if 'num_threads' in sp: 
        try: solver2.parameters.num_search_workers=int(sp['num_threads'])
        except: pass
    solver2.parameters.max_time_in_seconds = float(t2)
    try: solver2.parameters.relative_gap_limit = float(sp.get('relative_gap', 0.01))
    except: pass
    solver2.parameters.log_search_progress = True
    solver2.parameters.log_to_stdout = False    # Capture solver progress into unified log
    try:
        solver2.log_callback = lambda line: logging.getLogger("scheduler").info("[phase2] %s", line.rstrip())
    except Exception:
        pass
    if seed is not None: solver2.parameters.random_seed = int(seed)
    logger.info("Phase-2 solve: time=%ss workers=%s rgap=%s seed=%s",
                solver2.parameters.max_time_in_seconds,
                getattr(solver2.parameters, "num_search_workers", None),
                getattr(solver2.parameters, "relative_gap_limit", None),
                seed)

    # Diverse pool collection
    run_cfg = case.get("run", {}) or {}
    L = int(run_cfg.get("L", 0) or 0)
    logger.info("Diversity threshold L=%d, K=%s", L, K)

    cb = AssignmentPoolCollector(
        ctx2['x'], ctx2['S'], ctx2['P'],
        ctx2['days'], ctx2['providers'], ctx2['shifts'],
        sense='min', obj_slack=None, pool_limit=20000, dedup=True
    )
    st2 = solver2.Solve(model2, cb)
    logger.info("Phase-2 status=%s", solver2.StatusName())
    logger.info("Phase-2 best objective=%s best bound=%s", solver2.ObjectiveValue(), solver2.BestObjectiveBound())
    logger.info("Pool collected=%d", len(cb.pool))

    # Choose K diverse-best by Hamming ≥ L (with relaxation)
    selected_idx = _select_diverse_k(cb.pool, cb.pool_vecs, K, L, sense='min', relax_to=0)
    tables = [cb.pool[k][1] for k in selected_idx]
    per_meta = [cb.pool[k][2] for k in selected_idx]

    logger.info("Selected tables=%d (requested K=%d)", len(tables), K)
    if per_meta:
        logger.info("Selected objectives: %s", [m["objective"] for m in per_meta])

    meta2 = {
        "status": int(st2),
        "status_name": solver2.StatusName(),
        "response": solver2.ResponseStats(),
        "solutions_collected": len(cb.pool),
        "solutions_selected": len(tables),
        "best_objective": solver2.ObjectiveValue() if st2 in (cp_model.OPTIMAL, cp_model.FEASIBLE) else None,
        "best_bound": solver2.BestObjectiveBound() if st2 in (cp_model.OPTIMAL, cp_model.FEASIBLE) else None,
        "per_table": per_meta,
        "L": L
    }
    meta={"phase1": meta2, "phase2": meta2}
    return tables, meta

def write_excel_grid_multi(path, tables):
    wb=Workbook(); wb.remove(wb.active)
    for idx, table in enumerate(tables, start=1):
        days=table['days']; providers=table['providers']; shifts=table['shifts']; assign=set(table['assignment'])
        ws=wb.create_sheet(f"Schedule_{idx}")
        ws.cell(1,1,'Provider / Day')
        for j,d in enumerate(days, start=2): ws.cell(1,j,d)
        day_idx={d:k for k,d in enumerate(days)}
        grid=defaultdict(list)
        for (s, i) in assign:
            d=day_idx[shifts[s]['date']]
            grid[(i,d)].append(shifts[s]['type'])
        for i,prov in enumerate(providers, start=2):
            ws.cell(i,1,prov.get('name',f'Prov{i-1}'))
            for j,d in enumerate(days, start=2):
                ws.cell(i,j, ', '.join(sorted(grid.get((i-2,j-2), []))))

    if not tables:
        wb.create_sheet("Schedule_1")
    wb.save(path)

def write_excel_hospital_multi(path, tables):
    wb=Workbook(); wb.remove(wb.active)
    for idx, table in enumerate(tables, start=1):
        days=table['days']; providers=table['providers']; shifts=table['shifts']; assign=set(table['assignment'])
        ws=wb.create_sheet(f"Hospital_{idx}")
        ws.append(['Date','Role','Code','Start','End','Provider','ID'])
        for s,sh in enumerate(shifts):
            assignee='UNFILLED'
            for i in range(len(providers)):
                if (s, i) in assign: assignee=providers[i].get('name',f'Prov{i+1}'); break
            role,code=(sh['type'].split('_',1)+[''])[:2] if '_' in sh['type'] else ('', sh['type'])
            ws.append([sh['date'], role, code, sh['start'], sh['end'], assignee, sh.get('id', f'S{s:04d}')])
    if not tables:
        ws=wb.create_sheet("Hospital_1"); ws.append(['Date','Role','Code','Start','End','Provider','ID'])
    global CHOSPITAL
    CHOSPITAL = path
    wb.save(path)
    
def write_excel_calendar_multi(path, tables):
    """
    New calendar export:
      • One sheet per solution: "Calendar_<k>"
      • 8 columns: ["Shift", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
      • Organized in weekly blocks:
          - First row of each week shows the day numbers for Mon..Sun.
          - Under that, there are N rows (N = number of distinct shift TYPES
            appearing anywhere in that week). Left column lists those TYPES.
          - Each day's cell (for that type) contains a single NAME (if assigned).
            If multiple assignees exist, shows the first name + " (+N)".
            If that type exists that day but is unfilled, shows "UNFILLED".
            If that type doesn't exist that day, leaves the cell blank.
      • Weekend columns (Sat, Sun) shaded light red.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill, Font

    # Helpers
    def weekday_idx(date_str):  # Monday=0..Sunday=6
        y, m, d = map(int, date_str.split('-'))
        import datetime as _dt
        return _dt.date(y, m, d).weekday()

    def day_num(date_str):
        return int(date_str.split('-')[2])

    def group_days_into_weeks(days):
        """Return list of weeks; each week is a list len=7 of date_str or None, mon..sun."""
        if not days:
            return []
        first = days[0]
        first_wd = weekday_idx(first)  # 0..6 (Mon..Sun)
        weeks = []
        week = [None] * 7
        col = first_wd
        for d in days:
            week[col] = d
            col += 1
            if col >= 7:
                weeks.append(week)
                week = [None] * 7
                col = 0
        if any(week):
            weeks.append(week)
        return weeks

    def types_on_day(d, shifts):
        """Distinct shift TYPES present on a given date d."""
        return sorted({sh['type'] for sh in shifts if sh.get('date') == d})

    # Styles
    from openpyxl.styles import Alignment, PatternFill, Font
    wrap_mid = Alignment(horizontal="center", vertical="top", wrap_text=True)
    center_mid = Alignment(horizontal="center", vertical="center")
    left_mid = Alignment(horizontal="left", vertical="center")
    bold_font = Font(bold=True)
    wknd_fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")

    wb = Workbook()
    wb.remove(wb.active)

    for idx, table in enumerate(tables, start=1):
        days = table['days']
        providers = table['providers']
        shifts = table['shifts']
        assign_pairs = set(table['assignment'])

        # Build per-day TYPE -> [names]
        per_day = {}
        for s_idx, sh in enumerate(shifts):
            d = sh['date']
            t = sh['type']
            # find assignee(s)
            names = []
            for i in range(len(providers)):
                if (s_idx, i) in assign_pairs:
                    names.append(providers[i].get('name', f'Prov{i+1}'))
            per_day.setdefault(d, {}).setdefault(t, [])
            per_day[d][t].extend(names)

        # Build weeks Mon..Sun
        weeks = group_days_into_weeks(days)

        ws = wb.create_sheet(f"Calendar_{idx}")
        headers = ["Shift", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.font = bold_font
            cell.alignment = center_mid
        # Shade weekend headers
        ws.cell(row=1, column=7).fill = wknd_fill
        ws.cell(row=1, column=8).fill = wknd_fill

        # Column widths
        for col in range(1, 9):
            ws.column_dimensions[chr(ord('A') + col - 1)].width = 22 if col > 1 else 18

        r = 2  # next row to write
        for week in weeks:
            # Determine the set of shift TYPES appearing anywhere in this week
            week_types = set()
            for c, d in enumerate(week):  # c: 0..6 (Mon..Sun)
                if d is None:
                    continue
                week_types.update(types_on_day(d, shifts))
            week_types = sorted(week_types)

            # 1) Week "dates" header row
            ws.cell(row=r, column=1, value="")  # left-top cell blank
            for c in range(7):
                col = c + 2  # Mon..Sun => col 2..8
                d = week[c]
                val = int(d.split('-')[2]) if d else ""
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = bold_font
                cell.alignment = center_mid
                if col >= 8 or col == 7:  # Sat/Sun
                    cell.fill = wknd_fill
            r += 1

            # 2) Shift rows for this week
            if not week_types:
                r += 1
            else:
                for t in week_types:
                    left_cell = ws.cell(row=r, column=1, value=t)
                    left_cell.font = bold_font
                    left_cell.alignment = left_mid

                    for c in range(7):
                        col = c + 2
                        d = week[c]
                        cell = ws.cell(row=r, column=col)
                        if d is None:
                            val = ""
                        else:
                            day_types = per_day.get(d, {})
                            if t in day_types:
                                names = day_types[t]
                                if not names:
                                    val = "UNFILLED"
                                else:
                                    val = names[0] if len(names) == 1 else f"{names[0]} (+{len(names)-1})"
                            else:
                                val = ""
                        cell.value = val
                        cell.alignment = wrap_mid
                        if col >= 8 or col == 7:
                            cell.fill = wknd_fill
                    r += 1

            # Spacer
            r += 1

    if not tables:
        wb.create_sheet("Calendar_1")

    wb.save(path)



def compute_capacity_diag(case: Dict[str,Any]) -> List[Dict[str,Any]]:
    days = case['calendar']['days']
    shifts = case['shifts']
    providers = case['providers']
    provider_types = sorted(set([p.get('type','MD') for p in providers]))
    day_to_shifts = defaultdict(list)
    for s,sh in enumerate(shifts):
        day_to_shifts[sh['date']].append(s)
    out = []
    for i,p in enumerate(providers):
            ptype = p.get('type','MD')
        forb = set(p.get('forbidden_days_hard', []))
        ok_days = 0
        for d in days:
            if d in forb: continue
            any_ok = False
            for s in day_to_shifts[d]:
                if ptype in infer_allowed_types(shifts[s], provider_types):
                    any_ok = True; break
            if any_ok: ok_days += 1
        lim = p.get('limits', {}) or {}
        out.append({
            "provider": p.get('name', f'Prov{i+1}'),
            "eligible_days_upper_bound": ok_days,
            "min_total": lim.get('min_total', 0),
            "max_total": lim.get('max_total', None),
        })
    return out


def Solve_test_case(case):
    # Pre-init timestamp so logs & files share the same run id
    ts=dt.datetime.now().strftime('%Y%m%d_%H%M%S')

    # Lightweight run config probe (to derive out_dir for logger)
    try:
        with open(case, 'r', encoding='utf-8') as _f:
            _raw = json.load(_f)
        _run_cfg = (_raw.get("run") or {})
        out_dir = _run_cfg.get("out", "out")
    except Exception:
        out_dir = "out"

    logger = _mk_logger(out_dir, ts)

    # Redirect ALL stdout/stderr to the unified logger (file + console)
    sys.stdout = _StreamToLogger(logger, logging.INFO)
    sys.stderr = _StreamToLogger(logger, logging.ERROR)
    logger.info("===== SCHEDULER RUN %s =====", ts)
    logger.info("Args.case=%s", case)

    # Load merged inputs
    consts, case = load_inputs_from_case(case)
    logger.info("Loaded case with %d days, %d shifts, %d providers",
                len(case.get('calendar',{}).get('days',[])),
                len(case.get('shifts',[])),
                len(case.get('providers',[])))

    # Pull run config from the case
    run_cfg = case.get("run", {}) or {}
    out_dir = run_cfg.get("out", "out")
    K = int(run_cfg.get("k", 5) or 5)
    seed = run_cfg.get("seed", None)
    time_override = run_cfg.get("time", None)  # total time in seconds (overrides constants)
    L_cfg = int(run_cfg.get("L", 0) or 0)
    logger.info("Run config: out=%s k=%s seed=%s time=%s L=%s",
                out_dir, K, seed, time_override, L_cfg)

    # Ensure output dir exists
    os.makedirs(out_dir, exist_ok=True)

    # Effective-constants report (note reflects run-config override)
    _write_constants_log(consts, out_dir, cli_time_override=time_override)

    # Apply time override into constants if provided
    if time_override is not None:
        consts.setdefault('solver', {})['max_time_in_seconds'] = float(time_override)

    # Diagnostics snapshot (capacity bound)
    caps = compute_capacity_diag(case)
    caps_path = os.path.join(out_dir, 'eligibility_capacity.json')
    with open(caps_path, 'w', encoding='utf-8') as f:
        json.dump(caps, f, indent=2)
    logger.info("Wrote capacity snapshot: %s", caps_path)

    # Build & solve
    ctx = build_model(consts, case)
    logger.info("Model built: |S|=%d |P|=%d |D|=%d", len(ctx['S']), len(ctx['P']), len(ctx['D']))
    tables, meta = solve_two_phase(consts, case, ctx, K, seed=seed if seed is None else int(seed))

    # Outputs
    grid_path=os.path.join(out_dir, f'schedules.xlsx')
    hosp_path=os.path.join(out_dir, f'hospital_schedule.xlsx')
    cal_path=os.path.join(out_dir, f'calendar.xlsx')
    write_excel_grid_multi(grid_path, tables)
    write_excel_hospital_multi(hosp_path, tables)
    write_excel_calendar_multi(cal_path, tables)

    meta['run'] = {"timestamp": ts, "seed": seed, "out_dir": out_dir,
                   "files": {"grid": grid_path, "hospital": hosp_path, "calendar": cal_path,
                             "capacity": caps_path}}
    meta_path = os.path.join(out_dir, f'scheduler_log_{ts}.json')
    with open(meta_path,'w', encoding='utf-8') as f:
        json.dump(meta, f, indent=2)

    # Console echoes kept (now flow into unified log as well)
    print("Wrote:", grid_path)
    print("Wrote:", hosp_path)
    print("Wrote:", cal_path)
    print("Wrote:", os.path.join(out_dir,'eligibility_capacity.json'))

    # Additional logging
    logger.info("Wrote grid: %s", grid_path)
    logger.info("Wrote hospital: %s", hosp_path)
    logger.info("Wrote calendar: %s", cal_path)
    logger.info("Wrote run meta: %s", meta_path)
    logger.info("===== SCHEDULER RUN COMPLETE %s =====", ts)

    return tables, meta
# ---------- Defaults ----------
IDENTITY_MAX = 31  # "infinity" for limits and max_consecutive_days

DEFAULT_CONSTANTS = {
    "solver": {
        "max_time_in_seconds": 1000,
        "phase1_fraction": 0.4,
        "relative_gap": 0.00001,
        "num_threads": 8,
    },
    "weights": {
        "hard": {
            "slack_unfilled": 20,
            "slack_shift_less": 1,
            "slack_shift_more": 1,
            "slack_cant_work": 20,
            "slack_consec": 1,
        },  # could contain slack_name: BIG_WEIGHT to enforce as hard
        "soft": {
            "cluster": 10000,
            "cluster_size": 1,  # default requested
            "requested_off": 10000000,
            "days_wanted_not_met": 10000000,
            "cluster_weekend_start": 10000000,
            "unfair_number": 5000
        },
    },
    "objective": {"hard": 1, "soft": 1, "fair": 0},
}

# NOTE: L (variety constant) lives in Run tab, default here
DEFAULT_RUN = {"out": "", "k": 5, "L": 0, "seed": 123, "time": 180.0}

DEFAULT_CASE = {
    "constants": DEFAULT_CONSTANTS,
    "run": DEFAULT_RUN,
    "calendar": {"days": [], "weekend_days": ["Saturday", "Sunday"]},
    "shifts": [],
    "providers": []
}

WEEKDAY_NAMES = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
CANON_PROVIDER_TYPES = ["MD","DO","RN","NP","PA"]  # used by perturb tool
HARD_INF_WEIGHT = 1_000_000_000

# Colors for OFF/ON day visualization
OFF_FIXED_BG  = "#ffcdd2"  # light red
OFF_PREFER_BG = "#ffe0b2"  # light orange
ON_FIXED_BG   = "#c8e6c9"  # light green
ON_PREFER_BG  = "#bbdefb"  # light blue

# ---------- helpers ----------
def month_days(year:int, month:int):
    d = date(year, month, 1)
    res = []
    while d.month == month:
        res.append(d.isoformat()); d += timedelta(days=1)
    return res

def iso_dt(date_str, hhmm="08:00"):
    # store seconds as :00 internally; UI shows HH:MM
    if "T" in date_str:
        return date_str
    if len(hhmm) == 5:
        hhmm = hhmm + ":00"
    return f"{date_str}T{hhmm}"
def _next_day(date_str: str) -> str:
    y, m, d = map(int, date_str.split("-"))
    return (date(y, m, d) + timedelta(days=1)).isoformat()

def iso_dt_end(start_date: str, start_hm: str, end_hm: str) -> str:
    """
    Build an ISO datetime for the END. If end_hm < start_hm, roll to next day.
    """
    end_date = start_date if end_hm >= start_hm else _next_day(start_date)
    return iso_dt(end_date, end_hm)

def _normalize_overnight_shifts(shifts_list):
    """
    For already-loaded shifts: if end < start (as datetimes), bump end by +1 day.
    Safe no-op otherwise.
    """
    for sh in shifts_list:
        st = sh.get("start"); en = sh.get("end")
        try:
            dt_st = dt.datetime.fromisoformat(st)
            dt_en = dt.datetime.fromisoformat(en)
            if dt_en < dt_st:
                dt_en = dt_en + timedelta(days=1)
                sh["end"] = dt_en.strftime("%Y-%m-%dT%H:%M:%S")
        except Exception:
            # ignore malformed datetimes; UI edits will fix future ones
            pass

def unique_id(base, date_str, existing):
    cand = base if base not in existing else f"{base}@{date_str}"
    if cand not in existing: return cand
    k = 2
    while f"{cand}#{k}" in existing: k += 1
    return f"{cand}#{k}"

def now_out_name():
    return datetime.now().strftime("out_%Y%m%d_%H%M%S")

def _parse_int_default(s, default):
    try:
        return int(str(s).strip())
    except Exception:
        return default

def _sanitize_provider_identity_defaults(p):
    """Force identity defaults and remove nulls in provider constraints."""
    # max_consecutive_days
    v = p.get("max_consecutive_days", None)
    if not isinstance(v, int) or v <= 0:
        p["max_consecutive_days"] = IDENTITY_MAX

    # limits
    lim = p.setdefault("limits", {})
    lim["min_total"] = _parse_int_default(lim.get("min_total", 0), 0)
    mt = lim.get("max_total", IDENTITY_MAX)
    lim["max_total"] = _parse_int_default(mt, IDENTITY_MAX)

    # type_ranges: ensure no nulls, coerce
    tr = lim.get("type_ranges", {})
    if not isinstance(tr, dict):
        tr = {}
    fixed_tr = {}
    for t, rng in tr.items():
        mn, mx = 0, IDENTITY_MAX
        if isinstance(rng, (list, tuple)) and len(rng) == 2:
            mn = _parse_int_default(rng[0], 0)
            mx = _parse_int_default(rng[1], IDENTITY_MAX)
        fixed_tr[str(t)] = [mn, mx]
    lim["type_ranges"] = fixed_tr

    # Ensure list/dict fields aren't null
    p["forbidden_days_hard"] = list(p.get("forbidden_days_hard") or [])
    p["forbidden_days_soft"] = list(p.get("forbidden_days_soft") or [])
    p["preferred_days_hard"] = dict(p.get("preferred_days_hard") or {})
    p["preferred_days_soft"] = dict(p.get("preferred_days_soft") or {})

def _sanitize_case_no_nulls(case_dict):
    """Apply identity defaults & remove nulls for constraint-like fields."""
    for p in case_dict.get("providers", []):
        _sanitize_provider_identity_defaults(p)
    return case_dict

# ---------- main GUI ----------
class TestcaseGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Scheduler Testcase Builder")

        self._log_queue = queue.Queue()
        self._start_gui_log_pump()

        self.case = json.loads(json.dumps(DEFAULT_CASE))  # deep copy
        # default out folder = current datetime
        self.case["run"]["out"] = now_out_name()
        self.current_path = None

        # solver subprocess state
        self._proc = None
        self._io_queue = queue.Queue()
        self._io_thread = None

        # listbox default bgs (set after widgets are created)
        self._bg_off_default = None
        self._bg_on_default = None

        self._build_menu()
        self._build_tabs()
        self.try_autoload()
        self.refresh_all(select_first=True)

    # ---------- Menus ----------
    def _build_menu(self):
        menubar = tk.Menu(self.root)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="New", command=self.new_case)
        filemenu.add_command(label="Load...", command=self.load_case)
        self.sample_path = self._find_sample_path()
        if self.sample_path:
            filemenu.add_command(label=f"Load Sample ({Path(self.sample_path).name})",
                                 command=lambda: self.load_specific(self.sample_path))
        filemenu.add_command(label="Save", command=self.save_case)
        filemenu.add_command(label="Save As...", command=self.save_case_as)
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command=self.root.quit)

        toolsmenu = tk.Menu(menubar, tearoff=0)
        toolsmenu.add_command(label="Randomly Perturb Case…", command=self.menu_perturb_case)

        menubar.add_cascade(label="File", menu=filemenu)
        menubar.add_cascade(label="Tools", menu=toolsmenu)
        self.root.config(menu=menubar)

    def _find_sample_path(self):
        if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]): return sys.argv[1]
        for name in ["case_oct.json", "case_october_exact.json", "case.json", "sample_case.json"]:
            if os.path.isfile(name): return name
        return None

    def _on_solve_done(self):
        try:
            self.pb.stop()
            self.btn_run.configure(state="normal")
        except Exception:
            pass

        # ---------- Tabs ----------
    def _build_tabs(self):
        nb = ttk.Notebook(self.root)
        self.tab_run = ttk.Frame(nb)
        self.tab_cal = ttk.Frame(nb)
        self.tab_shifts = ttk.Frame(nb)
        self.tab_prov = ttk.Frame(nb)
        self.tab_cfg = ttk.Frame(nb)
        nb.add(self.tab_run, text="Run")
        nb.add(self.tab_cal, text="Calendar")
        nb.add(self.tab_shifts, text="Shifts")
        nb.add(self.tab_prov, text="Providers")
        nb.add(self.tab_cfg, text="Config")
        nb.pack(fill="both", expand=True)
        self._build_tab_run(self.tab_run)
        self._build_tab_calendar(self.tab_cal)
        self._build_tab_shifts(self.tab_shifts)
        self._build_tab_providers(self.tab_prov)
        self._build_tab_config(self.tab_cfg)

    # ---------- Run Tab ----------
    def _build_tab_run(self, frame):
        box = ttk.LabelFrame(frame, text="Run Settings")
        box.pack(fill="x", padx=8, pady=8)

        ttk.Label(box, text="Output folder name").grid(row=0, column=0, sticky="w", padx=6, pady=3)
        self.ent_out = ttk.Entry(box, width=30)
        self.ent_out.grid(row=0, column=1, sticky="w", padx=6, pady=3)
        ttk.Button(box, text="Open Output Folder…", command=self.open_out_folder).grid(row=0, column=2, padx=6)

        ttk.Label(box, text="k").grid(row=1, column=0, sticky="w", padx=6, pady=3)
        self.ent_r_k = ttk.Entry(box, width=8); self.ent_r_k.grid(row=1, column=1, sticky="w")

        # L (variety / min Hamming distance)
        ttk.Label(box, text="L (variety)").grid(row=2, column=0, sticky="w", padx=6, pady=3)
        self.ent_r_L = ttk.Entry(box, width=8); self.ent_r_L.grid(row=2, column=1, sticky="w")

        ttk.Label(box, text="seed").grid(row=3, column=0, sticky="w", padx=6, pady=3)
        self.ent_r_seed = ttk.Entry(box, width=10); self.ent_r_seed.grid(row=3, column=1, sticky="w")
        ttk.Label(box, text="time (min)").grid(row=4, column=0, sticky="w", padx=6, pady=3)
        self.ent_r_time = ttk.Entry(box, width=10); self.ent_r_time.grid(row=4, column=1, sticky="w")

        runbtns = ttk.Frame(frame); runbtns.pack(fill="x", padx=8, pady=(0,8))
        self.btn_run = ttk.Button(runbtns, text="Run Solver", command=self.run_solver)
        self.btn_run.pack(side="left", padx=(0,6))

        prog = ttk.Frame(frame); prog.pack(fill="x", padx=8, pady=(0,8))
        ttk.Label(prog, text="Progress").pack(anchor="w")
        self.pb = ttk.Progressbar(prog, orient="horizontal", mode="indeterminate")
        self.pb.pack(fill="x")

        logf = ttk.LabelFrame(frame, text="Log")
        logf.pack(fill="both", expand=True, padx=8, pady=8)
        self.txt_log = tk.Text(logf, height=16, wrap="word")
        self.txt_log.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(logf, orient="vertical", command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")

    # ---------- Calendar Tab ----------
    def _build_tab_calendar(self, frame):
        top = ttk.Frame(frame); top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text="Year").grid(row=0, column=0, sticky="w")
        self.ent_year = ttk.Entry(top, width=8); self.ent_year.grid(row=0, column=1, padx=5)
        ttk.Label(top, text="Month (1-12)").grid(row=0, column=2, sticky="w")
        self.ent_month = ttk.Entry(top, width=5); self.ent_month.grid(row=0, column=3, padx=5)
        ttk.Button(top, text="Generate Days", command=self.generate_days).grid(row=0, column=4, padx=10)

        wk = ttk.LabelFrame(frame, text="Weekend Days")
        wk.pack(fill="x", padx=8, pady=4)
        self.weekend_vars = {name: tk.BooleanVar(value=(name in self.case["calendar"]["weekend_days"])) for name in WEEKDAY_NAMES}
        row=0
        for i,name in enumerate(WEEKDAY_NAMES):
            cb = ttk.Checkbutton(wk, text=name, variable=self.weekend_vars[name], command=self.apply_weekends)
            cb.grid(row=row, column=i%4, sticky="w", padx=6, pady=4)
            if (i%4)==3: row+=1

        daysf = ttk.LabelFrame(frame, text="Days")
        daysf.pack(fill="both", expand=True, padx=8, pady=8)
        self.lst_days_cal = tk.Listbox(daysf, height=16, selectmode="extended", exportselection=False)
        self.lst_days_cal.pack(fill="both", expand=True)

    def generate_days(self):
        try:
            y = int(self.ent_year.get()); m = int(self.ent_month.get())
            if not (1 <= m <= 12): raise ValueError()
        except Exception:
            messagebox.showerror("Invalid", "Enter valid Year and Month (1-12)."); return
        old_days = set(self.case["calendar"]["days"])
        new_days = month_days(y, m)
        self.case["calendar"]["days"] = new_days
        new_set = set(new_days)
        for p in self.case.get("providers", []):
            for key in ("preferred_days_hard", "preferred_days_soft"):
                if isinstance(p.get(key), dict):
                    p[key] = {d: v for d, v in p[key].items() if d in new_set}
            for key in ("forbidden_days_hard", "forbidden_days_soft"):
                if isinstance(p.get(key), list):
                    p[key] = [d for d in p[key] if d in new_set]
        self._log(f"[calendar] generated {len(new_days)} days, cleared stale preferences.")
        self.refresh_all(select_first=True)

    def apply_weekends(self):
        selected = [name for name,var in self.weekend_vars.items() if var.get()] or ["Saturday","Sunday"]
        self.case["calendar"]["weekend_days"] = selected

    def refresh_days_cal(self):
        self.lst_days_cal.delete(0, tk.END)
        for d in self.case["calendar"]["days"]:
            self.lst_days_cal.insert(tk.END, d)

    # ---------- Shifts Tab ----------
    def _build_tab_shifts(self, frame):
        left = ttk.Frame(frame); left.pack(side="left", fill="y", padx=8, pady=8)
        mid = ttk.Frame(frame); mid.pack(side="left", fill="both", expand=True, padx=8, pady=8)
        right = ttk.LabelFrame(frame, text="Shift Editor"); right.pack(side="left", fill="y", padx=8, pady=8)

        ttk.Label(left, text="Select Date (optional for add-all)").pack(anchor="w")
        self.lst_dates_shifts = tk.Listbox(left, height=22, exportselection=False)
        self.lst_dates_shifts.pack(fill="y")
        self.lst_dates_shifts.bind("<<ListboxSelect>>", self.on_shift_date_select)

        cols = ("id","type","start","end","allowed_types")
        self.tree_shifts = ttk.Treeview(mid, columns=cols, show="headings", height=20)
        for c in cols:
            self.tree_shifts.heading(c, text=c)
            self.tree_shifts.column(c, width=120 if c!="allowed_types" else 180, stretch=True)
        self.tree_shifts.pack(fill="both", expand=True)
        self.tree_shifts.bind("<<TreeviewSelect>>", self.on_shift_row_select)

        # Editor fields
        self.ent_shift_id = ttk.Entry(right, width=28)
        self.ent_shift_type = ttk.Entry(right, width=28)
        # time pickers (HH:MM)
        timef1 = ttk.Frame(right); timef1.grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(timef1, text="Start (HH:MM)").pack(anchor="w")
        self.ent_shift_start = ttk.Entry(timef1, width=10)
        self.ent_shift_start.pack(anchor="w", pady=(0,4))
        timef2 = ttk.Frame(right); timef2.grid(row=1, column=0, columnspan=2, sticky="w")
        ttk.Label(timef2, text="End (HH:MM)").pack(anchor="w")
        self.ent_shift_end = ttk.Entry(timef2, width=10)
        self.ent_shift_end.pack(anchor="w", pady=(0,4))

        ttk.Label(right, text="ID (e.g., ER_AM)").grid(row=2, column=0, sticky="w", pady=(6,2))
        self.ent_shift_id.grid(row=2, column=1, sticky="w")
        ttk.Label(right, text="Type").grid(row=3, column=0, sticky="w")
        self.ent_shift_type.grid(row=3, column=1, sticky="w")
        ttk.Label(right, text="Allowed Provider Types (comma)").grid(row=4, column=0, sticky="w")
        self.ent_allowed_types = ttk.Entry(right, width=28)
        self.ent_allowed_types.grid(row=4, column=1, sticky="w")

        self.var_all_days_add = tk.BooleanVar(value=True)
        ttk.Checkbutton(right, text="Add across ALL calendar days", variable=self.var_all_days_add)\
            .grid(row=5, column=0, columnspan=2, sticky="w", pady=(6,2))

        self.var_all_days_update = tk.BooleanVar(value=False)
        ttk.Checkbutton(right, text="Update across ALL days (by Type)", variable=self.var_all_days_update)\
            .grid(row=6, column=0, columnspan=2, sticky="w")
        self.var_all_days_delete = tk.BooleanVar(value=False)
        ttk.Checkbutton(right, text="Delete across ALL days (by Type+ID base)", variable=self.var_all_days_delete)\
            .grid(row=7, column=0, columnspan=2, sticky="w")

        btnf = ttk.Frame(right); btnf.grid(row=8, column=0, columnspan=2, pady=(8,2))
        ttk.Button(btnf, text="Add", command=self.add_shift).pack(side="left", padx=4)
        ttk.Button(btnf, text="Update", command=self.update_shift).pack(side="left", padx=4)
        ttk.Button(btnf, text="Delete", command=self.delete_shift).pack(side="left", padx=4)

    def on_shift_date_select(self, event=None):
        self.refresh_shift_table()

    def shifts_for_date(self, date_str):
        return [sh for sh in self.case["shifts"] if sh.get("date")==date_str]

    def refresh_shift_table(self):
        sel = self.lst_dates_shifts.curselection()
        if not sel and self.case["calendar"]["days"]:
            self.lst_dates_shifts.selection_set(0)
            sel = (0,)
        for r in self.tree_shifts.get_children():
            self.tree_shifts.delete(r)
        if not sel: return
        d = self.lst_dates_shifts.get(sel[0])
        for sh in self.shifts_for_date(d):
            self.tree_shifts.insert("", "end", iid=sh["id"], values=(
                sh.get("id",""), sh.get("type",""),
                (sh.get("start","").split("T")[1][:5] if "T" in sh.get("start","") else sh.get("start","")[:5]),
                (sh.get("end","").split("T")[1][:5] if "T" in sh.get("end","") else sh.get("end","")[:5]),
                ",".join(sh.get("allowed_provider_types", []))
            ))

    def on_shift_row_select(self, event=None):
        sel = self.tree_shifts.selection()
        if not sel: return
        iid = sel[0]
        sh = next((x for x in self.case["shifts"] if x.get("id")==iid), None)
        if not sh: return
        self.ent_shift_id.delete(0, tk.END); self.ent_shift_id.insert(0, sh.get("id","").split("@")[0])
        self.ent_shift_type.delete(0, tk.END); self.ent_shift_type.insert(0, sh.get("type",""))
        st = sh.get("start",""); en = sh.get("end","")
        st = st.split("T")[1][:5] if "T" in st else st[:5]
        en = en.split("T")[1][:5] if "T" in en else en[:5]
        self.ent_shift_start.delete(0, tk.END); self.ent_shift_start.insert(0, st or "08:00")
        self.ent_shift_end.delete(0, tk.END); self.ent_shift_end.insert(0, en or "16:00")
        self.ent_allowed_types.delete(0, tk.END); self.ent_allowed_types.insert(0, ",".join(sh.get("allowed_provider_types", [])))

    def _validate_hhmm(self, s):
        try:
            if len(s) == 5 and s[2] == ":":
                h = int(s[:2]); m = int(s[3:])
                return 0 <= h < 24 and 0 <= m < 60
        except Exception:
            return False
        return False

    def add_shift(self):
        base_id = self.ent_shift_id.get().strip()
        if not base_id:
            messagebox.showerror("Missing ID", "Shift ID is required."); return
        st_hm = (self.ent_shift_start.get().strip() or "08:00")
        en_hm = (self.ent_shift_end.get().strip() or "16:00")
        if not (self._validate_hhmm(st_hm) and self._validate_hhmm(en_hm)):
            messagebox.showerror("Invalid time", "Use HH:MM (24h)."); return
        allowed = [t.strip() for t in self.ent_allowed_types.get().split(",") if t.strip()] or ["MD"]

        if self.var_all_days_add.get() or not self.lst_dates_shifts.curselection():
            dates = list(self.case["calendar"]["days"])
            if not dates:
                messagebox.showerror("No days", "Generate or load a calendar first."); return
        else:
            idx = self.lst_dates_shifts.curselection()[0]; dates = [self.lst_dates_shifts.get(idx)]

        sh_type = self.ent_shift_type.get().strip() or base_id
        for d in dates:
            for sh in self.shifts_for_date(d):
                if sh.get("type","") == sh_type:
                    messagebox.showerror("Duplicate type", f"Type '{sh_type}' already exists on {d}."); return

        if en_hm < st_hm:
            self._log(f"[shifts] '{base_id}' ends after midnight (end < start). Marked as cross-day.")

        existing = {s.get("id","") for s in self.case["shifts"]}
        for d in dates:
            sid = unique_id(base_id, d, existing); existing.add(sid)
            self.case["shifts"].append({
                "id": sid, "date": d, "type": sh_type,
                "start": iso_dt(d, st_hm), "end": iso_dt(d, en_hm),
                "allowed_provider_types": allowed
            })
        self.refresh_shift_table()
        self._log(f"[shifts] added {len(dates)} shift(s) of type {sh_type}.")
    def _start_gui_log_pump(self):
        """Drain the log queue at ~30 FPS; only place that touches Tk Text."""
        drained = 0
        try:
            while drained < 500:  # bound per tick so UI stays snappy
                msg = self._log_queue.get_nowait()
                self._log(msg)
                drained += 1
        except queue.Empty:
            pass
        # Re-arm pump
        self.root.after(33, self._start_gui_log_pump)

    def _solve_in_thread(self, case_path):
        logger = logging.getLogger("scheduler")
        gui_handler = TkQueueHandler(self)
        gui_handler.setLevel(logging.INFO)
        gui_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
        logger.addHandler(gui_handler)
        try:
            Solve_test_case(case_path)
        finally:
            logger.removeHandler(gui_handler)
            self.root.after(0, self._on_solve_done)

    def update_shift(self):
        sel = self.tree_shifts.selection()
        if not sel:
            messagebox.showerror("Select", "Pick a shift row first."); return
        iid = sel[0]
        sh = next((x for x in self.case["shifts"] if x.get("id")==iid), None)
        if not sh: return

        base_id = self.ent_shift_id.get().strip() or sh["id"].split("@")[0]
        d = sh["date"]
        new_type = self.ent_shift_type.get().strip() or sh["type"]
        st_hm = (self.ent_shift_start.get().strip() or "08:00")
        en_hm = (self.ent_shift_end.get().strip() or "16:00")
        if not (self._validate_hhmm(st_hm) and self._validate_hhmm(en_hm)):
            messagebox.showerror("Invalid time", "Use HH:MM (24h)."); return

        if self.var_all_days_update.get():
            dates = set(x["date"] for x in self.case["shifts"])
            for dd in dates:
                for s2 in self.shifts_for_date(dd):
                    if s2.get("id") != iid and s2.get("type") == new_type:
                        messagebox.showerror("Duplicate type", f"Type '{new_type}' already exists on {dd}."); return
        else:
            for s2 in self.shifts_for_date(d):
                if s2.get("id") != iid and s2.get("type") == new_type:
                    messagebox.showerror("Duplicate type", f"Type '{new_type}' already exists on {d}."); return

        if self.var_all_days_update.get():
            old_type = sh["type"]
            count = 0
            for obj in self.case["shifts"]:
                if obj["type"] == old_type:
                    obj["type"] = new_type
                    obj["start"] = iso_dt(obj["date"], st_hm)
                    obj["end"] = iso_dt(obj["date"], en_hm)
                    count += 1
            self._log(f"[shifts] updated {count} shifts (all days) from type '{old_type}' -> '{new_type}'.")
        else:
            new_id_base = base_id
            new_id = f"{new_id_base}@{d}" if "@" in sh["id"] or any(s.get("id")==new_id_base for s in self.case["shifts"] if s is not sh) else new_id_base
            if new_id != sh["id"] and any(s.get("id")==new_id for s in self.case["shifts"]):
                messagebox.showerror("Duplicate", f"ID '{new_id}' already exists."); return
            sh["id"] = new_id
            sh["type"] = new_type
            sh["start"] = iso_dt(d, st_hm)
            sh["end"]   = iso_dt_end(d, st_hm, en_hm)  # next-day if needed
            allowed = [t.strip() for t in self.ent_allowed_types.get().split(",") if t.strip()] or ["MD"]
            sh["allowed_provider_types"] = allowed
            self._log(f"[shifts] updated 1 shift ({new_id}).")

        self.refresh_shift_table()

    def delete_shift(self):
        sel = self.tree_shifts.selection()
        if not sel: return
        iid = sel[0]
        sh = next((x for x in self.case["shifts"] if x.get("id")==iid), None)
        if not sh: return
        d = sh["date"]; t = sh.get("type","")

        if self.var_all_days_delete.get():
            base = sh["id"].split("@")[0]
            before = len(self.case["shifts"])
            self.case["shifts"] = [s for s in self.case["shifts"]
                                   if not (s.get("type")==t or s.get("id","").split("@")[0]==base)]
            removed = before - len(self.case["shifts"])
            self._log(f"[shifts] deleted {removed} shifts across all days (type '{t}' or base '{base}').")
        else:
            self.case["shifts"] = [s for s in self.case["shifts"] if s.get("id")!=iid]
            self._log(f"[shifts] deleted 1 shift ({iid}).")

        remaining_types_on_day = {s.get("type","") for s in self.shifts_for_date(d)}
        if t and t not in remaining_types_on_day:
            for p in self.case.get("providers", []):
                changed = False
                for key in ("preferred_days_hard", "preferred_days_soft"):
                    mp = p.get(key) or {}
                    L = mp.get(d, None)
                    if isinstance(L, list) and t in L:
                        L2 = [x for x in L if x != t]
                        if L2:
                            mp[d] = L2; changed = True
                        else:
                            mp.pop(d, None); changed = True
                if changed:
                    self._log(f"[prefs] cleaned '{t}' from {p.get('name','?')} on {d} (no longer exists).")

        self.refresh_shift_table()
        self.render_provider_summary()
        self._recolor_off_on_days()

    # ---------- Providers Tab ----------
    def _build_tab_providers(self, frame):
        left = ttk.Frame(frame); left.pack(side="left", fill="y", padx=8, pady=8)
        mid = ttk.Frame(frame); mid.pack(side="left", fill="y", padx=8, pady=8)
        right = ttk.Frame(frame); right.pack(side="left", fill="both", expand=True, padx=8, pady=8)

        ttk.Label(left, text="Providers").pack(anchor="w")
        self.lst_providers = tk.Listbox(left, height=18, exportselection=False)
        self.lst_providers.pack(fill="y")
        self.lst_providers.bind("<<ListboxSelect>>", self.on_provider_select)

        ed = ttk.LabelFrame(left, text="Edit Provider")
        ed.pack(fill="x", pady=6)
        ttk.Label(ed, text="Name").grid(row=0, column=0, sticky="w")
        self.ent_pname = ttk.Entry(ed, width=22); self.ent_pname.grid(row=0, column=1)
        ttk.Label(ed, text="Type").grid(row=1, column=0, sticky="w")
        self.ent_ptype = ttk.Entry(ed, width=22); self.ent_ptype.grid(row=1, column=1)
        ttk.Label(ed, text="max_consecutive_days").grid(row=2, column=0, sticky="w")
        self.ent_pmaxc = ttk.Entry(ed, width=22); self.ent_pmaxc.grid(row=2, column=1)
        ttk.Label(ed, text="min_total").grid(row=3, column=0, sticky="w")
        self.ent_pmin = ttk.Entry(ed, width=22); self.ent_pmin.grid(row=3, column=1)
        ttk.Label(ed, text="max_total").grid(row=4, column=0, sticky="w")
        self.ent_pmax = ttk.Entry(ed, width=22); self.ent_pmax.grid(row=4, column=1)
        ttk.Button(ed, text="Add", command=self.add_provider).grid(row=5, column=0, pady=6)
        ttk.Button(ed, text="Update", command=self.update_provider).grid(row=5, column=1, pady=6)
        ttk.Button(ed, text="Delete", command=self.delete_provider).grid(row=6, column=0, columnspan=2, pady=6)

        offf = ttk.LabelFrame(mid, text="Choose Days (OFF)")
        offf.pack(fill="both", expand=True, pady=8)
        self.lst_days_off = tk.Listbox(offf, selectmode="extended", height=10, exportselection=False)
        self.lst_days_off.pack(fill="both", expand=True)
        off_btns = ttk.Frame(offf); off_btns.pack(pady=(6,2))
        ttk.Button(off_btns, text="Set FIXED OFF", command=self.apply_fixed_off_days).pack(side="left", padx=4)
        ttk.Button(off_btns, text="Set PREFER OFF", command=self.apply_prefer_off_days).pack(side="left", padx=4)
        ttk.Button(off_btns, text="Clear OFF", command=self.clear_off_days).pack(side="left", padx=4)
        self._legend_off = self._mk_legend(offf, [("Fixed OFF", OFF_FIXED_BG), ("Prefer OFF", OFF_PREFER_BG)])
        self._legend_off.pack(pady=(2,6), anchor="w")

        onf = ttk.LabelFrame(mid, text="Choose Days (ON)")
        onf.pack(fill="both", expand=True, pady=8)
        self.lst_days_on = tk.Listbox(onf, selectmode="extended", height=10, exportselection=False)
        self.lst_days_on.pack(fill="both", expand=True)
        on_btns = ttk.Frame(onf); on_btns.pack(pady=(6,2))
        ttk.Button(on_btns, text="Set FIXED ON", command=self.apply_pref_hard_days).pack(side="left", padx=4)
        ttk.Button(on_btns, text="Set PREFER ON", command=self.apply_pref_soft_days).pack(side="left", padx=4)
        ttk.Button(on_btns, text="Clear ON Prefs", command=self.action_clear_prefs).pack(side="left", padx=4)
        self._legend_on = self._mk_legend(onf, [("Fixed ON", ON_FIXED_BG), ("Prefer ON", ON_PREFER_BG)])
        self._legend_on.pack(pady=(2,6), anchor="w")

        pr = ttk.LabelFrame(right, text="ON Preferences - Shift Types")
        pr.pack(fill="both", expand=True)

        top = ttk.Frame(pr); top.pack(fill="x", pady=4)
        ttk.Label(top, text="Preview Date").grid(row=0, column=0, sticky="w")
        self.cmb_pref_date = ttk.Combobox(top, values=[], width=15, state="readonly")
        self.cmb_pref_date.grid(row=0, column=1, padx=6)
        self.cmb_pref_date.bind("<<ComboboxSelected>>", self.refresh_shift_boxes)

        ttk.Label(top, text="Shift Types (allowed for provider on date)").grid(row=0, column=2, padx=10)

        body = ttk.Frame(pr); body.pack(fill="both", expand=True)
        self.lb_shifts_on_day = tk.Listbox(body, selectmode="multiple", height=12, exportselection=False)
        self.lb_shifts_on_day.grid(row=0, column=0, padx=6, pady=6)

        sumf = ttk.LabelFrame(right, text="Provider Preferences Summary")
        sumf.pack(fill="both", expand=True, pady=(6,0))
        self.txt_pref_summary = tk.Text(sumf, width=60, height=16, wrap="word")
        self.txt_pref_summary.pack(side="left", fill="both", expand=True)
        sb2 = ttk.Scrollbar(sumf, orient="vertical", command=self.txt_pref_summary.yview)
        self.txt_pref_summary.configure(yscrollcommand=sb2.set)
        sb2.pack(side="right", fill="y")

        self._bg_off_default = self.lst_days_off.cget("background")
        self._bg_on_default  = self.lst_days_on.cget("background")

    def _mk_legend(self, parent, items):
        f = ttk.Frame(parent)
        for (label, color) in items:
            swatch = tk.Label(f, width=2, height=1, bg=color, relief="solid", bd=1)
            swatch.pack(side="left", padx=(0,4))
            ttk.Label(f, text=label).pack(side="left", padx=(0,10))
        return f

    # Provider tab helpers
    def refresh_providers(self):
        self.lst_providers.delete(0, tk.END)
        for p in self.case["providers"]:
            self.lst_providers.insert(tk.END, p.get("name","(no name)"))

    def selected_provider_index(self):
        sel = self.lst_providers.curselection()
        return (sel[0] if sel else None)

    def get_provider(self, idx=None):
        if idx is None:
            idx = self.selected_provider_index()
        if idx is None: return None
        if idx < 0 or idx >= len(self.case["providers"]): return None
        return self.case["providers"][idx]

    def _selected_days_off(self):
        return [self.lst_days_off.get(i) for i in self.lst_days_off.curselection()]

    def _selected_days_on(self):
        return [self.lst_days_on.get(i) for i in self.lst_days_on.curselection()]

    def add_provider(self):
        name = self.ent_pname.get().strip() or f"Prov{len(self.case['providers'])+1}"
        typ = self.ent_ptype.get().strip() or "MD"
        maxc = _parse_int_default(self.ent_pmaxc.get().strip(), IDENTITY_MAX)
        pmin = _parse_int_default(self.ent_pmin.get().strip(), 0)
        pmax = _parse_int_default(self.ent_pmax.get().strip(), IDENTITY_MAX)
        prov = {
            "name": name, "type": typ,
            "forbidden_days_hard": [],
            "forbidden_days_soft": [],
            "preferred_days_hard": {},
            "preferred_days_soft": {},
            "max_consecutive_days": maxc,
            "limits": {
                "min_total": pmin,
                "max_total": pmax,
                "type_ranges": {}
            }
        }
        _sanitize_provider_identity_defaults(prov)
        self.case["providers"].append(prov)
        self.refresh_providers()

    def update_provider(self):
        idx = self.selected_provider_index()
        if idx is None: return
        p = self.case["providers"][idx]
        p["name"] = self.ent_pname.get().strip() or p.get("name","")
        p["type"] = self.ent_ptype.get().strip() or p.get("type","MD")
        p["max_consecutive_days"] = _parse_int_default(self.ent_pmaxc.get().strip(), IDENTITY_MAX)
        lim = p.setdefault("limits", {})
        lim["min_total"] = _parse_int_default(self.ent_pmin.get().strip(), 0)
        lim["max_total"] = _parse_int_default(self.ent_pmax.get().strip(), IDENTITY_MAX)
        _sanitize_provider_identity_defaults(p)
        self.refresh_providers()
        self.render_provider_summary()
        self._recolor_off_on_days()

    def delete_provider(self):
        idx = self.selected_provider_index()
        if idx is None: return
        del self.case["providers"][idx]
        self.refresh_providers()
        self.render_provider_summary()
        self._recolor_off_on_days()

    def on_provider_select(self, event=None):
        p = self.get_provider()
        if not p: return
        _sanitize_provider_identity_defaults(p)
        self.ent_pname.delete(0, tk.END); self.ent_pname.insert(0, p.get("name",""))
        self.ent_ptype.delete(0, tk.END); self.ent_ptype.insert(0, p.get("type","MD"))
        self.ent_pmaxc.delete(0, tk.END); self.ent_pmaxc.insert(0, str(p.get("max_consecutive_days", IDENTITY_MAX)))
        lim = p.get("limits", {})
        self.ent_pmin.delete(0, tk.END); self.ent_pmin.insert(0, str(lim.get("min_total", 0)))
        self.ent_pmax.delete(0, tk.END); self.ent_pmax.insert(0, str(lim.get("max_total", IDENTITY_MAX)))
        self.refresh_off_boxes()
        self.refresh_provider_days()
        self.refresh_pref_ui_sources()
        self.render_provider_summary()
        self._recolor_off_on_days()

    def refresh_off_boxes(self):
        self.lst_days_off.delete(0, tk.END)
        for d in self.case["calendar"]["days"]:
            self.lst_days_off.insert(tk.END, d)

    def refresh_provider_days(self):
        self.lst_days_on.delete(0, tk.END)
        for d in self.case["calendar"]["days"]:
            self.lst_days_on.insert(tk.END, d)

    def apply_fixed_off_days(self):
        p = self.get_provider()
        if not p:
            messagebox.showerror("No provider", "Select a provider first.")
            return
        sel = self._selected_days_off()
        p["forbidden_days_hard"] = sorted(set(p.get("forbidden_days_hard", [])) | set(sel))
        p["forbidden_days_soft"] = [d for d in p.get("forbidden_days_soft", []) if d not in sel]
        self._log(f"[prefs] set FIXED OFF ({len(sel)} days) for {p.get('name','?')}")
        self.render_provider_summary()
        self._recolor_off_on_days()

    def apply_prefer_off_days(self):
        p = self.get_provider()
        if not p:
            messagebox.showerror("No provider", "Select a provider first.")
            return
        sel = self._selected_days_off()
        p["forbidden_days_soft"] = sorted(set(p.get("forbidden_days_soft", [])) | set(sel))
        p["forbidden_days_hard"] = [d for d in p.get("forbidden_days_hard", []) if d not in sel]
        self._log(f"[prefs] set PREFER OFF ({len(sel)} days) for {p.get('name','?')}")
        self.render_provider_summary()
        self._recolor_off_on_days()

    def clear_off_days(self):
        p = self.get_provider()
        if not p:
            self._log("[warn] Select a provider first.", warn=True); return
        sel = set(self._selected_days_off())
        if not sel:
            self._log("[warn] Pick one or more OFF days to clear.", warn=True); return
        p["forbidden_days_hard"] = [d for d in p.get("forbidden_days_hard", []) if d not in sel]
        p["forbidden_days_soft"] = [d for d in p.get("forbidden_days_soft", []) if d not in sel]
        self._log(f"[prefs] cleared OFF flags on {len(sel)} day(s) for {p.get('name','?')}.")
        self.render_provider_summary()
        self._recolor_off_on_days()

    # ON prefs helpers
    def shift_types_by_date_filtered(self, d):
        p = self.get_provider()
        if not p or not d:
            return []
        ptype = p.get("type","MD")
        types = set()
        for sh in self.case["shifts"]:
            if sh.get("date") == d:
                allowed = set(sh.get("allowed_provider_types", ["MD"]))
                if ptype in allowed:
                    t = sh.get("type","")
                    if t: types.add(t)
        return sorted(types)

    def refresh_shift_boxes(self, event=None):
        self.lb_shifts_on_day.delete(0, tk.END)
        d = self.cmb_pref_date.get()
        if not d: return
        for t in self.shift_types_by_date_filtered(d):
            self.lb_shifts_on_day.insert(tk.END, t)

    def refresh_pref_ui_sources(self):
        days = self.case["calendar"]["days"]
        self.cmb_pref_date["values"] = days
        if days:
            self.cmb_pref_date.current(0)
        self.refresh_shift_boxes()

    def render_provider_summary(self):
        self.txt_pref_summary.delete("1.0", tk.END)
        p = self.get_provider()
        if not p:
            return
        _sanitize_provider_identity_defaults(p)
        lines = []
            lines.append(f"Provider: {p.get('name','?')}  (type={p.get('type','MD')})")
        lim = p.get("limits", {})
        lines.append(f"Limits: min_total={lim.get('min_total',0)} max_total={lim.get('max_total',IDENTITY_MAX)}")
        lines.append(f"max_consecutive_days: {p.get('max_consecutive_days', IDENTITY_MAX)}")
        lines.append("")
        lines.append("Fixed OFF days:")
        for d in p.get("forbidden_days_hard", []):
            lines.append(f"  - {d}")
        lines.append("Prefer OFF days:")
        for d in p.get("forbidden_days_soft", []):
            lines.append(f"  - {d}")
        lines.append("")
        lines.append("Fixed ON (date → types):")
        for d, L in sorted((p.get("preferred_days_hard") or {}).items()):
            lines.append(f"  - {d}: {', '.join(L) if L else '(none)'}")
        lines[-1] = lines[-1].replace("', '", "', '")
        lines.append("Prefer ON (date → types):")
        for d, L in sorted((p.get("preferred_days_soft") or {}).items()):
            lines.append(f"  - {d}: {', '.join(L) if L else '(none)'}")
        self.txt_pref_summary.insert(tk.END, "\n".join(lines))

    def _selected_shift_types_from_listbox(self):
        return [self.lb_shifts_on_day.get(i) for i in self.lb_shifts_on_day.curselection()]

    def apply_pref_hard_days(self):
        p = self.get_provider()
        days = self._selected_days_on()
        if not p or not days:
            self._log("[warn] Pick a provider and one or more ON days.", warn=True)
            return
        chosen = self._selected_shift_types_from_listbox()
        hard = p.setdefault("preferred_days_hard", {})
        for d in days:
            allowed_today = set(self.shift_types_by_date_filtered(d))
            val = sorted(t for t in chosen if t in allowed_today)
            hard[d] = val
        self._log(f"[prefs] set FIXED ON on {len(days)} day(s) for {p.get('name','?')}: {', '.join(chosen) if chosen else '(none)'}")
        self.render_provider_summary()
        self._recolor_off_on_days()

    def apply_pref_soft_days(self):
        p = self.get_provider()
        days = self._selected_days_on()
        if not p or not days:
            self._log("[warn] Pick a provider and one or more ON days.", warn=True)
            return
        chosen = self._selected_shift_types_from_listbox()
        soft = p.setdefault("preferred_days_soft", {})
        for d in days:
            allowed_today = set(self.shift_types_by_date_filtered(d))
            val = sorted(t for t in chosen if t in allowed_today)
            soft[d] = val
        self._log(f"[prefs] set PREFER ON on {len(days)} day(s) for {p.get('name','?')}: {', '.join(chosen) if chosen else '(none)'}")
        self.render_provider_summary()
        self._recolor_off_on_days()

    def action_clear_prefs(self):
        p = self.get_provider()
        days = self._selected_days_on()
        if not p or not days:
            self._log("[warn] Pick a provider and one or more ON days.", warn=True)
            return
        hard_map = p.setdefault("preferred_days_hard", {})
        soft_map = p.setdefault("preferred_days_soft", {})
        for d in days:
            hard_map.pop(d, None)
            soft_map.pop(d, None)
        self._log(f"[prefs] cleared ON prefs on {len(days)} day(s) for {p.get('name','?')}.")
        self.render_provider_summary()
        self._recolor_off_on_days()

    def _recolor_off_on_days(self):
        def reset_colors(lst, default_bg):
            try:
                n = lst.size()
                for i in range(n):
                    lst.itemconfig(i, bg=default_bg)
            except Exception:
                pass

        if self._bg_off_default is not None:
            reset_colors(self.lst_days_off, self._bg_off_default)
        if self._bg_on_default is not None:
            reset_colors(self.lst_days_on, self._bg_on_default)

        p = self.get_provider()
        if not p:
            return

        days = self.case["calendar"]["days"]

        fixed_off = set(p.get("forbidden_days_hard", []))
        prefer_off = set(p.get("forbidden_days_soft", []))
        fixed_on = set(d for d, L in (p.get("preferred_days_hard") or {}).items() if L)
        prefer_on = set(d for d, L in (p.get("preferred_days_soft") or {}).items() if L)

        for i, d in enumerate(days):
            try:
                if d in fixed_off:
                    self.lst_days_off.itemconfig(i, bg=OFF_FIXED_BG)
                elif d in prefer_off:
                    self.lst_days_off.itemconfig(i, bg=OFF_PREFER_BG)
            except Exception:
                pass

        for i, d in enumerate(days):
            try:
                if d in fixed_on:
                    self.lst_days_on.itemconfig(i, bg=ON_FIXED_BG)
                elif d in prefer_on:
                    self.lst_days_on.itemconfig(i, bg=ON_PREFER_BG)
            except Exception:
                pass

    # ---------- Config Tab ----------
    def _build_tab_config(self, frame):
        sol = ttk.LabelFrame(frame, text="constants.solver")
        sol.pack(fill="x", padx=8, pady=8)
        self.ent_s_time = ttk.Entry(sol, width=10)
        self.ent_s_phase = ttk.Entry(sol, width=10)
        self.ent_s_gap = ttk.Entry(sol, width=10)
        self.ent_s_threads = ttk.Entry(sol, width=10)
        row = 0
        for lbl, ent, hint in [
            ("max_time_in_seconds", self.ent_s_time, "e.g., 350"),
            ("phase1_fraction", self.ent_s_phase, "e.g., 0.4"),
            ("relative_gap", self.ent_s_gap, "e.g., 0.01"),
            ("num_threads", self.ent_s_threads, "e.g., 8"),
        ]:
            ttk.Label(sol, text=lbl).grid(row=row, column=0, sticky="w", padx=6, pady=3)
            ent.grid(row=row, column=1, sticky="w", padx=6, pady=3)
            ttk.Label(sol, text=hint, foreground="#666").grid(row=row, column=2, sticky="w")
            row += 1

        run = ttk.LabelFrame(frame, text="run (defaults shown in Run tab too)")
        run.pack(fill="x", padx=8, pady=8)
        self.ent_cfg_out = ttk.Entry(run, width=18)
        self.ent_cfg_k   = ttk.Entry(run, width=8)
        self.ent_cfg_seed= ttk.Entry(run, width=12)
        self.ent_cfg_time= ttk.Entry(run, width=10)
        for r,(lbl, ent) in enumerate([
            ("out", self.ent_cfg_out), ("k", self.ent_cfg_k),
            ("seed", self.ent_cfg_seed), ("time", self.ent_cfg_time)
        ]):
            ttk.Label(run, text=lbl).grid(row=r, column=0, sticky="w", padx=6, pady=3)
            ent.grid(row=r, column=1, sticky="w", padx=6, pady=3)

        adv = ttk.LabelFrame(frame, text="Advanced (raw JSON)")
        adv.pack(fill="both", expand=True, padx=8, pady=8)
        ttk.Label(adv, text="constants.weights").grid(row=0, column=0, sticky="w")
        ttk.Label(adv, text="constants.objective").grid(row=0, column=1, sticky="w")
        self.txt_weights = tk.Text(adv, width=48, height=12)
        self.txt_objective = tk.Text(adv, width=32, height=12)
        self.txt_weights.grid(row=1, column=0, padx=6, pady=6, sticky="nsew")
        self.txt_objective.grid(row=1, column=1, padx=6, pady=6, sticky="nsew")

        adv.grid_columnconfigure(0, weight=1)
        adv.grid_columnconfigure(1, weight=1)
        adv.grid_rowconfigure(1, weight=1)

        ttk.Button(frame, text="Apply Config to Case", command=self.apply_config).pack(pady=8)

    def load_config_into_ui(self):
        c = self.case.get("constants", {}) or {}
        s = c.get("solver", {}) or {}
        w = c.get("weights", {}) or {}
        o = c.get("objective", {}) or {}
        r = self.case.get("run", {}) or {}

        self.ent_s_time.delete(0, tk.END);   self.ent_s_time.insert(0, str(s.get("max_time_in_seconds","")))
        self.ent_s_phase.delete(0, tk.END);  self.ent_s_phase.insert(0, str(s.get("phase1_fraction","")))
        self.ent_s_gap.delete(0, tk.END);    self.ent_s_gap.insert(0, str(s.get("relative_gap","")))
        self.ent_s_threads.delete(0, tk.END);self.ent_s_threads.insert(0, str(s.get("num_threads","")))

        self.ent_cfg_out.delete(0, tk.END); self.ent_cfg_out.insert(0, r.get("out",""))
        self.ent_cfg_k.delete(0, tk.END);   self.ent_cfg_k.insert(0, str(r.get("k","")))
        self.ent_cfg_seed.delete(0, tk.END); self.ent_cfg_seed.insert(0, "" if r.get("seed",None) is None else str(r.get("seed")))
        self.ent_cfg_time.delete(0, tk.END); self.ent_cfg_time.insert(0, str(r.get("time","")))

        self.ent_out.delete(0, tk.END); self.ent_out.insert(0, r.get("out",""))
        self.ent_r_k.delete(0, tk.END); self.ent_r_k.insert(0, str(r.get("k","")))
        self.ent_r_L.delete(0, tk.END); self.ent_r_L.insert(0, str(r.get("L","")))
        self.ent_r_seed.delete(0, tk.END); self.ent_r_seed.insert(0, "" if r.get("seed",None) is None else str(r.get("seed")))
        self.ent_r_time.delete(0, tk.END); self.ent_r_time.insert(0, str(r.get("time","")))

        self.txt_weights.delete("1.0", tk.END); self.txt_weights.insert(tk.END, json.dumps(w, indent=2))
        self.txt_objective.delete("1.0", tk.END); self.txt_objective.insert(tk.END, json.dumps(o, indent=2))

    def apply_config(self):
        def _parse_json_editor(widget, label):
            txt = widget.get("1.0", "end").strip()
            if not txt:
                return None
            try:
                obj = json.loads(txt)
            except json.JSONDecodeError as e:
                raise ValueError(f"{label} JSON is invalid:\n{e}")
            if not isinstance(obj, dict):
                raise ValueError(f"{label} must be a JSON object.")
            return obj

        try:
            s = self.case.setdefault("constants", {}).setdefault("solver", {})
            if self.ent_s_time.get().strip() != "":   s["max_time_in_seconds"] = float(self.ent_s_time.get().strip())
            if self.ent_s_phase.get().strip() != "":  s["phase1_fraction"]     = float(self.ent_s_phase.get().strip())
            if self.ent_s_gap.get().strip() != "":    s["relative_gap"]        = float(self.ent_s_gap.get().strip())
            if self.ent_s_threads.get().strip() != "":s["num_threads"]         = int(self.ent_s_threads.get().strip())

            consts = self.case.setdefault("constants", {})

            parsed_w = _parse_json_editor(self.txt_weights, "constants.weights")
            if parsed_w is not None:
                existing_w = consts.get("weights", {})
                merged_w = dict(existing_w); merged_w.update(parsed_w)
                consts["weights"] = merged_w

            parsed_o = _parse_json_editor(self.txt_objective, "constants.objective")
            if parsed_o is not None:
                existing_o = consts.get("objective", {})
                merged_o = dict(existing_o); merged_o.update(parsed_o)
                consts["objective"] = merged_o

            r = self.case.setdefault("run", {})
            if self.ent_cfg_out.get().strip() != "":   r["out"]  = self.ent_cfg_out.get().strip()
            if self.ent_cfg_k.get().strip() != "":     r["k"]    = int(self.ent_cfg_k.get().strip())
            seed_txt = self.ent_cfg_seed.get().strip()
            r["seed"] = None if seed_txt=="" else int(seed_txt)
            if self.ent_cfg_time.get().strip() != "":  r["time"] = float(self.ent_cfg_time.get().strip())

            self.load_config_into_ui()
            self._log("[config] Applied to case.")
        except Exception as e:
            messagebox.showerror("Invalid Config", f"Failed to apply config:\n{e}")

    # ---------- File ops ----------
    def new_case(self):
        self.case = json.loads(json.dumps(DEFAULT_CASE))
        self.case["run"]["out"] = now_out_name()
        self.current_path = None
        self.refresh_all(select_first=True)

    def _normalize_loaded_case(self, case_dict):
        consts = case_dict.setdefault("constants", {})
        consts.setdefault("solver", DEFAULT_CONSTANTS["solver"])
        weights = consts.setdefault("weights", {})
        hard = weights.setdefault("hard", {})
        hard.setdefault("slack_unfilled", 1)
        hard.setdefault("slack_shift_less", 1)
        hard.setdefault("slack_shift_more", 1)
        hard.setdefault("slack_cant_work", 1)
        hard.setdefault("slack_consec", 1)
        soft = weights.setdefault("soft", {})
        soft.setdefault("cluster", DEFAULT_CONSTANTS["weights"]["soft"]["cluster"])
        soft.setdefault("cluster_size", DEFAULT_CONSTANTS["weights"]["soft"]["cluster_size"])
        soft.setdefault("requested_off", DEFAULT_CONSTANTS["weights"]["soft"]["requested_off"])
        soft.setdefault("days_wanted_not_met", DEFAULT_CONSTANTS["weights"]["soft"]["days_wanted_not_met"])
        soft.setdefault("cluster_weekend_start", DEFAULT_CONSTANTS["weights"]["soft"]["cluster_weekend_start"])
        soft.setdefault("unfair_number", DEFAULT_CONSTANTS["weights"]["soft"]["unfair_number"])
        consts.setdefault("objective", DEFAULT_CONSTANTS["objective"])

        run = case_dict.setdefault("run", {})
        run.setdefault("out", now_out_name())
        run.setdefault("k", DEFAULT_RUN["k"])
        run.setdefault("L", DEFAULT_RUN["L"])
        run.setdefault("seed", DEFAULT_RUN["seed"])
        run.setdefault("time", DEFAULT_RUN["time"])

        case_dict.setdefault("calendar", {})
        case_dict["calendar"].setdefault("days", [])
        case_dict["calendar"].setdefault("weekend_days", ["Saturday","Sunday"])
        case_dict.setdefault("shifts", [])
        case_dict.setdefault("providers", [])
        for p in case_dict["providers"]:
            if isinstance(p.get("preferred_days_soft"), list):
                p["preferred_days_soft"] = {}
            if isinstance(p.get("preferred_days_hard"), list):
                p["preferred_days_hard"] = {}
            p.setdefault("forbidden_days_hard", [])
            p.setdefault("forbidden_days_soft", [])
            # Identity defaults (no nulls)
            _sanitize_provider_identity_defaults(p)
        _normalize_overnight_shifts(case_dict["shifts"])
        return case_dict

    def _load_from_path(self, fp):
        with open(fp, "r", encoding="utf-8") as f:
            raw = json.load(f)
        self.case = self._normalize_loaded_case(raw)
        self.current_path = fp
        self.refresh_all(select_first=True)

    def load_case(self):
        fp = filedialog.askopenfilename(title="Load case JSON", filetypes=[("JSON","*.json")])
        if not fp: return
        try:
            self._load_from_path(fp)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load:\n{e}")

    def load_specific(self, fp):
        try:
            self._load_from_path(fp)
            messagebox.showinfo("Loaded", f"Loaded {fp}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load:\n{e}")

    def try_autoload(self):
        fp = self._find_sample_path()
        if fp and os.path.isfile(fp):
            try:
                self._load_from_path(fp)
            except Exception:
                pass

    def _validate_case(self):
        errors = []
        days = set(self.case["calendar"]["days"])
        for sh in self.case["shifts"]:
            if sh.get("date") not in days:
                errors.append(f"Shift {sh.get('id')} has date not in calendar: {sh.get('date')}")
            for key in ("id","type","start","end"):
                if key not in sh:
                    errors.append(f"Shift missing field '{key}': {sh}")
            if "allowed_provider_types" not in sh:
                errors.append(f"Shift {sh.get('id')} missing allowed_provider_types")
        for p in self.case["providers"]:
            if "name" not in p:
                errors.append("Provider missing name")
            # Enforce identity defaults
            _sanitize_provider_identity_defaults(p)
        if errors:
            for e in errors:
                self._log(f"[validate] {e}", warn=True)
            return False
        self._log("[validate] case OK")
        return True

    def save_case(self):
        if not self.current_path:
            return self.save_case_as()
        try:
            _sanitize_case_no_nulls(self.case)
            if not self._validate_case():
                if not messagebox.askyesno("Validation", "Case has issues. Save anyway?"):
                    return
            with open(self.current_path, "w", encoding="utf-8") as f:
                json.dump(self.case, f, indent=2)
            self._log(f"[save] wrote {self.current_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save:\n{e}")

    def save_case_as(self):
        fp = filedialog.asksaveasfilename(title="Save case JSON", defaultextension=".json", filetypes=[("JSON","*.json")])
        if not fp: return
        self.current_path = fp
        self.save_case()

    # ---------- Tools: Random Perturbation ----------
    def menu_perturb_case(self):
        dlg = PerturbDialog(self.root, title="Randomly Perturb Case")
        if not getattr(dlg, "result", None):
            return
        r = dlg.result
        self.randomly_perturb_case(r["pct_providers"]/100.0, r["pct_shifts"]/100.0, r["seed"], r["tweak_weekend"])

    def randomly_perturb_case(self, frac_prov=0.25, frac_shift=0.25, seed=None, tweak_weekend=True):
        rng = random.Random(seed)
        days = list(self.case.get("calendar", {}).get("days", []))
        if not days:
            messagebox.showerror("No calendar", "Generate or load a calendar first.")
            return

        all_shift_types = sorted({sh.get("type","") for sh in self.case.get("shifts", [])} - {""})
        all_provider_types = sorted({p.get("type","MD") for p in self.case.get("providers", [])} | set(CANON_PROVIDER_TYPES))

        provs = self.case.get("providers", [])
        n_p = len(provs)
        k_p = max(1, int(round(n_p * frac_prov))) if n_p else 0
        prov_idx = rng.sample(range(n_p), k_p) if k_p and n_p else []
        for idx in prov_idx:
            p = provs[idx]
            if rng.random() < 0.5 and all_provider_types:
                p["type"] = rng.choice(all_provider_types)
            if rng.random() < 0.4:
                # Only numeric choices (no None)
                p["max_consecutive_days"] = rng.choice([rng.randint(4, 14), rng.randint(15, 22), IDENTITY_MAX])
            if rng.random() < 0.6:
                forb = set(p.get("forbidden_days_hard", []))
                n_flip = rng.randint(0, max(1, len(days)//10))
                for _ in range(n_flip):
                    d = rng.choice(days)
                    if d in forb: forb.remove(d)
                    else: forb.add(d)
                p["forbidden_days_hard"] = sorted(forb)
            if rng.random() < 0.6:
                soft = set(p.get("forbidden_days_soft", []))
                n_flip = rng.randint(0, max(1, len(days)//12))
                for _ in range(n_flip):
                    d = rng.choice(days)
                    if d in soft: soft.remove(d)
                    else: soft.add(d)
                p["forbidden_days_soft"] = sorted(soft)
            for key, prob in (("preferred_days_hard", 0.5), ("preferred_days_soft", 0.7)):
                if rng.random() < prob:
                    prefs = p.get(key) or {}
                    p[key] = prefs
                    pick = min(len(days), rng.randint(1, 4))
                    for d in rng.sample(days, pick):
                        if rng.random() < 0.25:
                            prefs[d] = []
                        else:
                            if all_shift_types:
                                n_types = rng.randint(1, min(3, max(1, len(all_shift_types))))
                                prefs[d] = sorted(rng.sample(all_shift_types, n_types))
            _sanitize_provider_identity_defaults(p)

        shifts = self.case.get("shifts", [])
        n_sh = len(shifts)
        k_sh = max(1, int(round(n_sh * frac_shift))) if n_sh else 0
        sh_idx = rng.sample(range(n_sh), k_sh) if k_sh and n_sh else []

        time_choices = ["06:00","07:00","08:00","12:00","16:00","18:00","20:00","22:00"]

        for idx in sh_idx:
            sh = shifts[idx]
            d = sh.get("date", rng.choice(days))
            if rng.random() < 0.25:
                try:
                    base = datetime.strptime(d, "%Y-%m-%d").date()
                    cand = [base + timedelta(days=off) for off in rng.sample([-1,1], k=1)]
                    cand = [c for c in cand if c.isoformat() in days]
                    if cand:
                        d = cand[0].isoformat()
                        sh["date"] = d
                except Exception:
                    pass
            if rng.random() < 0.8 and all_shift_types:
                sh["type"] = rng.choice(all_shift_types)
            if rng.random() < 0.85:
                start_choice = rng.choice(time_choices[:-2])
                later_choices = [t for t in time_choices if t > start_choice]
                end_choice = rng.choice(later_choices) if later_choices else "23:00"
                sh["start"] = iso_dt(d, start_choice)
                sh["end"]   = iso_dt(d, end_choice)
            if rng.random() < 0.85:
                n_ap = rng.randint(1, min(3, len(CANON_PROVIDER_TYPES)))
                sh["allowed_provider_types"] = sorted(random.sample(CANON_PROVIDER_TYPES, n_ap))

        self.refresh_all(select_first=True)
        self._log(f"[perturb] Changed providers ~{int(frac_prov*100)}% and shifts ~{int(frac_shift*100)}%.")

    # ---------- Run solver ----------
    def open_out_folder(self):
        out = self.ent_out.get().strip()
        if not out:
            out = now_out_name()
            self.ent_out.delete(0, tk.END); self.ent_out.insert(0, out)
        p = Path(out)
        if not p.exists():
            p.mkdir(parents=True, exist_ok=True)
        if sys.platform.startswith("win"):
            os.startfile(str(p))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(p)])
        else:
            subprocess.Popen(["xdg-open", str(p)])

    def run_solver(self):
        r = self.case.setdefault("run", {})
        out = self.ent_out.get().strip() or now_out_name()
        self.ent_out.delete(0, tk.END); self.ent_out.insert(0, out)
        r["out"] = out
        try:
            r["k"] = int(self.ent_r_k.get().strip() or r.get("k", 5))
            L_txt = self.ent_r_L.get().strip()
            r["L"] = int(L_txt) if L_txt != "" else int(r.get("L", 0))
            seed_txt = self.ent_r_seed.get().strip()
            r["seed"] = None if seed_txt=="" else int(seed_txt)
            time_minutes = float(self.ent_r_time.get().strip() or r.get("time", 180.0))
            r["time"] = time_minutes
        except Exception as e:
            messagebox.showerror("Run", f"Bad run parameters: {e}")
            return

        solver = self.case.setdefault("constants", {}).setdefault("solver", {})
        phase1 = float(solver.get("phase1_fraction", 0.0) or 0.0)
        effective_seconds = float(r["time"]) * 60.0 * (1.0 + phase1)
        solver["max_time_in_seconds"] = effective_seconds

        odir = Path(out)
        if odir.exists():
            if any(odir.iterdir()):
                self._warn_to_log(f"[run] Output folder '{out}' is not empty. Choose another name.")
                return
        else:
            odir.mkdir(parents=True, exist_ok=True)


        case_path = (odir / "case.json").resolve()
        case_path = os.path.abspath(case_path)  # before Solve_test_case(...)

        with open(case_path, "w", encoding="utf-8") as f:
            json.dump(_sanitize_case_no_nulls(self.case), f, indent=2)
            f.flush()
            os.fsync(f.fileno())  # make sure file is on disk before spawning

        self._log(f"=== RUN @ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
        self._log(f"[run] scheduler_sat.py → {odir}. k={r['k']}  L={r.get('L',0)}  time_min={r['time']}")

        self.btn_run.configure(state="disabled")
        self.pb.start(12)

        this_script = os.path.abspath(__file__)

        self._proc = subprocess.Popen(
            [sys.executable, this_script, "--_solver_child", "--case", str(case_path)],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1
        )

        self._io_thread = threading.Thread(target=self._pump_proc_output, daemon=True)
        self._io_thread.start()
        self._drain_log_queue()

        # Start child; capture stdout+stderr line-buffered
        self._proc = subprocess.Popen(
            [sys.executable, this_script, "--_solver_child", "--case", str(case_path)],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,           # universal_newlines
            bufsize=1            # line-buffered
        )

        # Pump the child's output on a background thread into our internal queue…
        self._io_thread = threading.Thread(target=self._pump_proc_output, daemon=True)
        self._io_thread.start()
        # …and start draining that queue onto the Text widget from the Tk loop
        self._drain_log_queue()

    def cancel_solver(self):
        if self._proc and (self._proc.poll() is None):
            try:
                self._proc.terminate()
            except Exception:
                pass
            self._log("[run] cancel requested…")
        else:
            self._warn_to_log("[no active run]")

    def _pump_proc_output(self):
        try:
            for line in self._proc.stdout:
                self._io_queue.put(line)
        except Exception:
            pass
        finally:
            self._io_queue.put(None)

    def _drain_log_queue(self):
        try:
            while True:
                item = self._io_queue.get_nowait()
                if item is None:
                    rc = self._proc.poll() if self._proc else None
                    self._log(f"[done] solver exited with code {rc}")
                    self.pb.stop()
                    self.btn_run.configure(state="normal")
                    self._proc = None
                    return
                self._log(item.rstrip("\n"))
        except queue.Empty:
            pass
        self.root.after(100, self._drain_log_queue)

    # ---------- Refresh ----------
    def refresh_all(self, select_first=False):
        self.refresh_days_cal()
        self.lst_dates_shifts.delete(0, tk.END)
        for d in self.case["calendar"]["days"]:
            self.lst_dates_shifts.insert(tk.END, d)
        if select_first and self.case["calendar"]["days"]:
            self.lst_dates_shifts.selection_clear(0, tk.END)
            self.lst_dates_shifts.selection_set(0)
        self.refresh_shift_table()

        self.refresh_providers()
        self.refresh_off_boxes()
        self.refresh_provider_days()
        self.refresh_pref_ui_sources()

        self.load_config_into_ui()
        self._recolor_off_on_days()

    # ---------- Logging ----------
    def _log(self, msg, warn=False):
        if warn:
            msg = f"[WARN] {msg}"
        self.txt_log.insert(tk.END, msg + "\n")
        self.txt_log.see(tk.END)

    def _warn_to_log(self, msg):
        self._log(msg, warn=True)

# ---------- Perturb dialog ----------
class PerturbDialog(simpledialog.Dialog):
    def body(self, master):
        ttk.Label(master, text="Percent of providers to change (5–100)").grid(row=0, column=0, sticky="w")
        self.var_p_prov = tk.IntVar(value=25)
        self.scale_prov = tk.Scale(master, from_=5, to=100, orient="horizontal", variable=self.var_p_prov)
        self.scale_prov.grid(row=0, column=1, sticky="ew", padx=6)

        ttk.Label(master, text="Percent of shifts to change (5–100)").grid(row=1, column=0, sticky="w")
        self.var_p_shift = tk.IntVar(value=25)
        self.scale_shift = tk.Scale(master, from_=5, to=100, orient="horizontal", variable=self.var_p_shift)
        self.scale_shift.grid(row=1, column=1, sticky="ew", padx=6)

        ttk.Label(master, text="Random seed (optional)").grid(row=2, column=0, sticky="w")
        self.entry_seed = ttk.Entry(master, width=16)
        self.entry_seed.grid(row=2, column=1, sticky="w")

        self.var_tweak_weekend = tk.BooleanVar(value=True)
        ttk.Checkbutton(master, text="Allow tweaking weekend days", variable=self.var_tweak_weekend)\
            .grid(row=3, column=0, columnspan=2, sticky="w", pady=(6,0))

        master.grid_columnconfigure(1, weight=1)
        return self.entry_seed

    def apply(self):
        seed_text = self.entry_seed.get().strip()
        self.result = {
            "pct_providers": max(5, min(100, self.var_p_prov.get())),
            "pct_shifts":    max(5, min(100, self.var_p_shift.get())),
            "seed":          None if seed_text == "" else seed_text,
            "tweak_weekend": self.var_tweak_weekend.get(),
        }

# ---------- solver-child entry (no Tk) ----------
def _solver_child_main(argv):
    # Expect: --_solver_child --case <path>
    try:
        i = argv.index("--case")
        case_path = argv[i + 1]
    except Exception:
        print("[solver-child] Missing --case <path>")
        sys.exit(2)

    # Run the solver (this sets up logging redirection inside the child)
    Solve_test_case(case_path)
    global CHOSPITAL




    print(case_path, CHOSPITAL)

    run_diag(case_path, CHOSPITAL)

# ---------- main ----------
def main():
    root = tk.Tk()
    gui = TestcaseGUI(root)
    root.minsize(1200, 800)
    root.mainloop()

if __name__ == "__main__":
    if "--_solver_child" in sys.argv:
        _solver_child_main(sys.argv)
    else:
        main()